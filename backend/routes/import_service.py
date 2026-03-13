# OCB TITAN ERP - Import Service
# Validator and import engine for master data
# tenant-aware, atomic transactions, error reporting

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from datetime import datetime, timezone
from typing import List, Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import uuid
from database import get_db
from utils.auth import get_current_user
from models.titan_models import AuditLog

router = APIRouter(prefix="/import", tags=["Import"])


class ImportResult:
    def __init__(self):
        self.success_count = 0
        self.error_count = 0
        self.errors = []
        self.imported_ids = []


def validate_required_field(value, field_name: str, row_num: int) -> Optional[str]:
    """Validate required field is not empty"""
    if value is None or str(value).strip() == "":
        return f"Baris {row_num}: {field_name} wajib diisi"
    return None


def validate_unique(existing_values: set, value: str, field_name: str, row_num: int) -> Optional[str]:
    """Validate value is unique"""
    if value in existing_values:
        return f"Baris {row_num}: {field_name} '{value}' sudah ada (duplikat)"
    return None


def validate_numeric(value, field_name: str, row_num: int, min_val: float = 0) -> Optional[str]:
    """Validate numeric value"""
    try:
        num = float(value) if value else 0
        if num < min_val:
            return f"Baris {row_num}: {field_name} tidak boleh kurang dari {min_val}"
        return None
    except (ValueError, TypeError):
        return f"Baris {row_num}: {field_name} harus berupa angka"


# ==================== PRODUCTS IMPORT ====================

@router.post("/products")
async def import_products(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """
    Import Products from Excel
    
    Required columns: Kode, Nama, Harga Beli, Harga Jual
    Optional columns: Kategori, Satuan, Min Stok, Deskripsi
    
    Validation:
    - Kode (SKU) harus unik
    - Harga harus >= 0
    - Harga Jual >= Harga Beli
    """
    if user.get("role") not in ["owner", "admin", "warehouse"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus format Excel (.xlsx/.xls)")
    
    db = get_db()
    result = ImportResult()
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        # Get existing SKUs
        existing_products = await db["products"].find({}, {"_id": 0, "code": 1}).to_list(100000)
        existing_skus = {p["code"].lower() for p in existing_products if p.get("code")}
        
        # Get categories for validation
        categories = await db["categories"].find({"is_active": True}, {"_id": 0}).to_list(1000)
        category_map = {c.get("name", "").lower(): c for c in categories}
        
        # Get units
        units = await db["units"].find({"is_active": True}, {"_id": 0}).to_list(1000)
        unit_map = {u.get("name", "").lower(): u for u in units}
        
        # Find header row
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(cell and "kode" in str(cell).lower() for cell in row):
                header_row = row_idx
                break
        
        if not header_row:
            raise HTTPException(status_code=400, detail="Header tidak ditemukan. Pastikan ada kolom 'Kode'")
        
        # Map column indices
        headers = [str(cell).lower().strip() if cell else "" for cell in ws[header_row]]
        col_map = {
            "code": next((i for i, h in enumerate(headers) if "kode" in h or "sku" in h), None),
            "name": next((i for i, h in enumerate(headers) if "nama" in h), None),
            "cost_price": next((i for i, h in enumerate(headers) if "beli" in h or "cost" in h), None),
            "selling_price": next((i for i, h in enumerate(headers) if "jual" in h or "sell" in h), None),
            "category": next((i for i, h in enumerate(headers) if "kategori" in h or "category" in h), None),
            "unit": next((i for i, h in enumerate(headers) if "satuan" in h or "unit" in h), None),
            "min_stock": next((i for i, h in enumerate(headers) if "min" in h and "stok" in h), None),
            "description": next((i for i, h in enumerate(headers) if "desk" in h or "desc" in h), None),
        }
        
        if col_map["code"] is None or col_map["name"] is None:
            raise HTTPException(status_code=400, detail="Kolom 'Kode' dan 'Nama' wajib ada")
        
        # Process data rows
        products_to_insert = []
        new_skus = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue
            
            # Extract values
            code = str(row[col_map["code"]]).strip() if row[col_map["code"]] else ""
            name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
            cost_price = row[col_map["cost_price"]] if col_map["cost_price"] is not None and row[col_map["cost_price"]] else 0
            selling_price = row[col_map["selling_price"]] if col_map["selling_price"] is not None and row[col_map["selling_price"]] else 0
            category_name = str(row[col_map["category"]]).strip() if col_map["category"] is not None and row[col_map["category"]] else ""
            unit_name = str(row[col_map["unit"]]).strip() if col_map["unit"] is not None and row[col_map["unit"]] else ""
            min_stock = row[col_map["min_stock"]] if col_map["min_stock"] is not None and row[col_map["min_stock"]] else 0
            description = str(row[col_map["description"]]).strip() if col_map["description"] is not None and row[col_map["description"]] else ""
            
            # Validate
            errors = []
            
            err = validate_required_field(code, "Kode", row_idx)
            if err:
                errors.append(err)
            
            err = validate_required_field(name, "Nama", row_idx)
            if err:
                errors.append(err)
            
            if code:
                err = validate_unique(existing_skus, code.lower(), "Kode/SKU", row_idx)
                if err:
                    errors.append(err)
                err = validate_unique(new_skus, code.lower(), "Kode/SKU (dalam file)", row_idx)
                if err:
                    errors.append(err)
            
            err = validate_numeric(cost_price, "Harga Beli", row_idx)
            if err:
                errors.append(err)
            
            err = validate_numeric(selling_price, "Harga Jual", row_idx)
            if err:
                errors.append(err)
            
            try:
                if float(selling_price or 0) < float(cost_price or 0):
                    errors.append(f"Baris {row_idx}: Harga Jual tidak boleh kurang dari Harga Beli")
            except:
                pass
            
            if errors:
                result.error_count += 1
                result.errors.extend(errors)
                continue
            
            # Build product document
            category = category_map.get(category_name.lower())
            unit = unit_map.get(unit_name.lower())
            
            product = {
                "id": str(uuid.uuid4()),
                "code": code,
                "name": name,
                "cost_price": float(cost_price or 0),
                "selling_price": float(selling_price or 0),
                "category_id": category.get("id") if category else None,
                "category_name": category.get("name") if category else category_name,
                "unit_id": unit.get("id") if unit else None,
                "unit_name": unit.get("name") if unit else unit_name or "Pcs",
                "min_stock": int(float(min_stock or 0)),
                "description": description,
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user.get("user_id", ""),
                "import_source": file.filename
            }
            
            products_to_insert.append(product)
            new_skus.add(code.lower())
            result.imported_ids.append(product["id"])
        
        # Atomic insert
        if products_to_insert:
            await db["products"].insert_many(products_to_insert)
            result.success_count = len(products_to_insert)
            
            # Audit log
            audit = AuditLog(
                user_id=user.get("user_id", ""),
                user_name=user.get("name", ""),
                action="import",
                module="products",
                entity_type="product",
                entity_id=f"batch_{len(products_to_insert)}",
                new_value={
                    "count": len(products_to_insert),
                    "source_file": file.filename
                }
            )
            await db["audit_logs"].insert_one(audit.model_dump())
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
    
    return {
        "message": f"Import selesai: {result.success_count} berhasil, {result.error_count} gagal",
        "success_count": result.success_count,
        "error_count": result.error_count,
        "errors": result.errors[:50],  # First 50 errors
        "imported_ids": result.imported_ids[:20]  # First 20 IDs
    }


# ==================== SUPPLIERS IMPORT ====================

@router.post("/suppliers")
async def import_suppliers(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """
    Import Suppliers from Excel
    
    Required columns: Kode, Nama
    Optional columns: Telepon, Email, Alamat, NPWP, Kontak
    
    Validation:
    - Kode harus unik
    """
    if user.get("role") not in ["owner", "admin"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus format Excel (.xlsx/.xls)")
    
    db = get_db()
    result = ImportResult()
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        existing = await db["suppliers"].find({}, {"_id": 0, "code": 1}).to_list(100000)
        existing_codes = {s["code"].lower() for s in existing if s.get("code")}
        
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(cell and "kode" in str(cell).lower() for cell in row):
                header_row = row_idx
                break
        
        if not header_row:
            raise HTTPException(status_code=400, detail="Header tidak ditemukan")
        
        headers = [str(cell).lower().strip() if cell else "" for cell in ws[header_row]]
        col_map = {
            "code": next((i for i, h in enumerate(headers) if "kode" in h), None),
            "name": next((i for i, h in enumerate(headers) if "nama" in h), None),
            "phone": next((i for i, h in enumerate(headers) if "telp" in h or "phone" in h or "hp" in h), None),
            "email": next((i for i, h in enumerate(headers) if "email" in h), None),
            "address": next((i for i, h in enumerate(headers) if "alamat" in h or "address" in h), None),
            "npwp": next((i for i, h in enumerate(headers) if "npwp" in h), None),
            "contact_person": next((i for i, h in enumerate(headers) if "kontak" in h or "contact" in h or "pic" in h), None),
        }
        
        if col_map["code"] is None or col_map["name"] is None:
            raise HTTPException(status_code=400, detail="Kolom 'Kode' dan 'Nama' wajib ada")
        
        suppliers_to_insert = []
        new_codes = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue
            
            code = str(row[col_map["code"]]).strip() if row[col_map["code"]] else ""
            name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
            
            errors = []
            
            err = validate_required_field(code, "Kode", row_idx)
            if err:
                errors.append(err)
            
            err = validate_required_field(name, "Nama", row_idx)
            if err:
                errors.append(err)
            
            if code:
                err = validate_unique(existing_codes, code.lower(), "Kode", row_idx)
                if err:
                    errors.append(err)
                err = validate_unique(new_codes, code.lower(), "Kode (dalam file)", row_idx)
                if err:
                    errors.append(err)
            
            if errors:
                result.error_count += 1
                result.errors.extend(errors)
                continue
            
            supplier = {
                "id": str(uuid.uuid4()),
                "code": code,
                "name": name,
                "phone": str(row[col_map["phone"]]).strip() if col_map["phone"] is not None and row[col_map["phone"]] else "",
                "email": str(row[col_map["email"]]).strip() if col_map["email"] is not None and row[col_map["email"]] else "",
                "address": str(row[col_map["address"]]).strip() if col_map["address"] is not None and row[col_map["address"]] else "",
                "npwp": str(row[col_map["npwp"]]).strip() if col_map["npwp"] is not None and row[col_map["npwp"]] else "",
                "contact_person": str(row[col_map["contact_person"]]).strip() if col_map["contact_person"] is not None and row[col_map["contact_person"]] else "",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user.get("user_id", ""),
                "import_source": file.filename
            }
            
            suppliers_to_insert.append(supplier)
            new_codes.add(code.lower())
            result.imported_ids.append(supplier["id"])
        
        if suppliers_to_insert:
            await db["suppliers"].insert_many(suppliers_to_insert)
            result.success_count = len(suppliers_to_insert)
            
            audit = AuditLog(
                user_id=user.get("user_id", ""),
                user_name=user.get("name", ""),
                action="import",
                module="suppliers",
                entity_type="supplier",
                entity_id=f"batch_{len(suppliers_to_insert)}",
                new_value={"count": len(suppliers_to_insert), "source_file": file.filename}
            )
            await db["audit_logs"].insert_one(audit.model_dump())
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
    
    return {
        "message": f"Import selesai: {result.success_count} berhasil, {result.error_count} gagal",
        "success_count": result.success_count,
        "error_count": result.error_count,
        "errors": result.errors[:50],
        "imported_ids": result.imported_ids[:20]
    }


# ==================== CUSTOMERS IMPORT ====================

@router.post("/customers")
async def import_customers(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """
    Import Customers from Excel
    
    Required columns: Nama
    Optional columns: Kode, Telepon, Email, Alamat, NPWP, Grup
    
    Validation:
    - Kode harus unik (jika ada)
    """
    if user.get("role") not in ["owner", "admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Insufficient permissions")
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File harus format Excel (.xlsx/.xls)")
    
    db = get_db()
    result = ImportResult()
    
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(BytesIO(content))
        ws = wb.active
        
        existing = await db["customers"].find({}, {"_id": 0, "code": 1}).to_list(100000)
        existing_codes = {c["code"].lower() for c in existing if c.get("code")}
        
        # Get customer groups
        groups = await db["customer_groups"].find({"is_active": True}, {"_id": 0}).to_list(100)
        group_map = {g.get("name", "").lower(): g for g in groups}
        
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
            if row and any(cell and ("nama" in str(cell).lower() or "name" in str(cell).lower()) for cell in row):
                header_row = row_idx
                break
        
        if not header_row:
            raise HTTPException(status_code=400, detail="Header tidak ditemukan")
        
        headers = [str(cell).lower().strip() if cell else "" for cell in ws[header_row]]
        col_map = {
            "code": next((i for i, h in enumerate(headers) if "kode" in h), None),
            "name": next((i for i, h in enumerate(headers) if "nama" in h or "name" in h), None),
            "phone": next((i for i, h in enumerate(headers) if "telp" in h or "phone" in h or "hp" in h), None),
            "email": next((i for i, h in enumerate(headers) if "email" in h), None),
            "address": next((i for i, h in enumerate(headers) if "alamat" in h or "address" in h), None),
            "npwp": next((i for i, h in enumerate(headers) if "npwp" in h), None),
            "group": next((i for i, h in enumerate(headers) if "grup" in h or "group" in h), None),
        }
        
        if col_map["name"] is None:
            raise HTTPException(status_code=400, detail="Kolom 'Nama' wajib ada")
        
        customers_to_insert = []
        new_codes = set()
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue
            
            name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
            code = str(row[col_map["code"]]).strip() if col_map["code"] is not None and row[col_map["code"]] else ""
            group_name = str(row[col_map["group"]]).strip() if col_map["group"] is not None and row[col_map["group"]] else ""
            
            errors = []
            
            err = validate_required_field(name, "Nama", row_idx)
            if err:
                errors.append(err)
            
            if code:
                err = validate_unique(existing_codes, code.lower(), "Kode", row_idx)
                if err:
                    errors.append(err)
                err = validate_unique(new_codes, code.lower(), "Kode (dalam file)", row_idx)
                if err:
                    errors.append(err)
            
            if errors:
                result.error_count += 1
                result.errors.extend(errors)
                continue
            
            # Auto-generate code if not provided
            if not code:
                code = f"CUS-{str(uuid.uuid4())[:8].upper()}"
            
            group = group_map.get(group_name.lower())
            
            customer = {
                "id": str(uuid.uuid4()),
                "code": code,
                "name": name,
                "phone": str(row[col_map["phone"]]).strip() if col_map["phone"] is not None and row[col_map["phone"]] else "",
                "email": str(row[col_map["email"]]).strip() if col_map["email"] is not None and row[col_map["email"]] else "",
                "address": str(row[col_map["address"]]).strip() if col_map["address"] is not None and row[col_map["address"]] else "",
                "npwp": str(row[col_map["npwp"]]).strip() if col_map["npwp"] is not None and row[col_map["npwp"]] else "",
                "group_id": group.get("id") if group else None,
                "group_name": group.get("name") if group else group_name,
                "is_active": True,
                "points": 0,
                "total_purchase": 0,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "created_by": user.get("user_id", ""),
                "import_source": file.filename
            }
            
            customers_to_insert.append(customer)
            new_codes.add(code.lower())
            result.imported_ids.append(customer["id"])
        
        if customers_to_insert:
            await db["customers"].insert_many(customers_to_insert)
            result.success_count = len(customers_to_insert)
            
            audit = AuditLog(
                user_id=user.get("user_id", ""),
                user_name=user.get("name", ""),
                action="import",
                module="customers",
                entity_type="customer",
                entity_id=f"batch_{len(customers_to_insert)}",
                new_value={"count": len(customers_to_insert), "source_file": file.filename}
            )
            await db["audit_logs"].insert_one(audit.model_dump())
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
    
    return {
        "message": f"Import selesai: {result.success_count} berhasil, {result.error_count} gagal",
        "success_count": result.success_count,
        "error_count": result.error_count,
        "errors": result.errors[:50],
        "imported_ids": result.imported_ids[:20]
    }


# ==================== IMPORT TEMPLATES ====================

@router.get("/template/{type}")
async def get_import_template(
    type: str,
    user: dict = Depends(get_current_user)
):
    """Download import template for products, suppliers, or customers"""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    
    templates = {
        "products": {
            "title": "Template Import Produk",
            "headers": ["Kode/SKU*", "Nama*", "Harga Beli*", "Harga Jual*", "Kategori", "Satuan", "Min Stok", "Deskripsi"],
            "sample": ["PRD-001", "Produk Contoh", 10000, 15000, "Aksesoris", "Pcs", 5, "Deskripsi produk"]
        },
        "suppliers": {
            "title": "Template Import Supplier",
            "headers": ["Kode*", "Nama*", "Telepon", "Email", "Alamat", "NPWP", "Kontak Person"],
            "sample": ["SUP-001", "Supplier Contoh", "08123456789", "supplier@email.com", "Jl. Contoh No. 1", "", "John Doe"]
        },
        "customers": {
            "title": "Template Import Customer",
            "headers": ["Kode", "Nama*", "Telepon", "Email", "Alamat", "NPWP", "Grup"],
            "sample": ["CUS-001", "Customer Contoh", "08987654321", "customer@email.com", "Jl. Customer No. 1", "", "Retail"]
        }
    }
    
    if type not in templates:
        raise HTTPException(status_code=400, detail="Template type invalid. Use: products, suppliers, customers")
    
    template = templates[type]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template["title"][:31]
    
    # Headers
    for col, header in enumerate(template["headers"], 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Sample data
    for col, value in enumerate(template["sample"], 1):
        ws.cell(row=2, column=col, value=value)
    
    # Auto-width
    for col in range(1, len(template["headers"]) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"template_import_{type}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
