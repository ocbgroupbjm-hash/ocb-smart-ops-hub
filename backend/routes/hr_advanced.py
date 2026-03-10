# OCB TITAN AI - Advanced HR System
# Complete HR with relations, documents, training, structure

from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from typing import List, Optional
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
import uuid
import os
import aiofiles

from database import get_db

router = APIRouter(prefix="/api/hr", tags=["HR System"])

UPLOAD_DIR = "/app/uploads"
os.makedirs(f"{UPLOAD_DIR}/employees", exist_ok=True)
os.makedirs(f"{UPLOAD_DIR}/documents", exist_ok=True)
os.makedirs(f"{UPLOAD_DIR}/attendance", exist_ok=True)

def gen_id():
    return str(uuid.uuid4())

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def get_wib_now():
    return datetime.now(timezone.utc) + timedelta(hours=7)

# Collections
def employees_col():
    return get_db()['employees']

def departments_col():
    return get_db()['departments']

def positions_col():
    return get_db()['master_jabatan']

def branches_col():
    return get_db()['branches']

def employee_docs_col():
    return get_db()['employee_documents']

def trainings_col():
    return get_db()['trainings']

def training_participants_col():
    return get_db()['training_participants']

def org_structure_col():
    return get_db()['organization_structure']

def contracts_col():
    return get_db()['employee_contracts']

def promotions_col():
    return get_db()['employee_promotions']

def warnings_col():
    return get_db()['employee_warnings']

# ==================== MODELS ====================

class Department(BaseModel):
    code: str
    name: str
    parent_id: str = ""
    head_employee_id: str = ""
    description: str = ""

class EmployeeRelation(BaseModel):
    employee_id: str
    relation_type: str  # supervisor, subordinate, mentor, mentee
    related_employee_id: str

class EmployeeDocument(BaseModel):
    employee_id: str
    document_type: str  # ktp, npwp, ijazah, sertifikat, kontrak, sk
    document_name: str
    expiry_date: str = ""
    notes: str = ""

class Training(BaseModel):
    title: str
    description: str = ""
    trainer: str
    training_type: str  # internal, external, online
    start_date: str
    end_date: str
    location: str = ""
    max_participants: int = 50
    department_ids: List[str] = []
    position_ids: List[str] = []
    is_mandatory: bool = False

class EmployeeContract(BaseModel):
    employee_id: str
    contract_type: str  # tetap, kontrak, magang, freelance
    start_date: str
    end_date: str
    salary: float
    notes: str = ""

class EmployeeWarning(BaseModel):
    employee_id: str
    warning_type: str  # sp1, sp2, sp3, verbal
    reason: str
    issued_date: str
    issued_by: str

class EmployeePromotion(BaseModel):
    employee_id: str
    from_position_id: str
    to_position_id: str
    from_salary: float
    to_salary: float
    effective_date: str
    reason: str = ""
    approved_by: str = ""

# ==================== DEPARTMENTS ====================

@router.get("/departments")
async def get_departments():
    """Get all departments with hierarchy"""
    depts = await departments_col().find({}, {"_id": 0}).to_list(length=100)
    
    # Build hierarchy
    dept_map = {d.get("id"): d for d in depts}
    for d in depts:
        parent_id = d.get("parent_id")
        if parent_id and parent_id in dept_map:
            d["parent_name"] = dept_map[parent_id].get("name")
        
        # Get head employee
        if d.get("head_employee_id"):
            head = await employees_col().find_one({"id": d["head_employee_id"]}, {"_id": 0, "name": 1})
            d["head_name"] = head.get("name") if head else ""
        
        # Count employees
        count = await employees_col().count_documents({"department": d.get("name")})
        d["employee_count"] = count
    
    return {"departments": depts}

@router.post("/departments")
async def create_department(dept: Department):
    """Create department"""
    doc = {
        "id": gen_id(),
        **dept.dict(),
        "is_active": True,
        "created_at": now_iso()
    }
    await departments_col().insert_one(doc)
    doc.pop("_id", None)  # Remove MongoDB _id before returning
    return {"message": "Department berhasil dibuat", "department": doc}

@router.put("/departments/{dept_id}")
async def update_department(dept_id: str, dept: Department):
    """Update department"""
    await departments_col().update_one(
        {"id": dept_id},
        {"$set": {**dept.dict(), "updated_at": now_iso()}}
    )
    return {"message": "Department berhasil diupdate"}

# ==================== EMPLOYEE RELATIONS ====================

@router.get("/employees/{employee_id}/relations")
async def get_employee_relations(employee_id: str):
    """Get employee relations (supervisor, subordinates, etc.)"""
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    # Get supervisor
    supervisor = None
    if emp.get("supervisor_id"):
        supervisor = await employees_col().find_one({"id": emp["supervisor_id"]}, {"_id": 0, "id": 1, "name": 1, "jabatan_name": 1})
    
    # Get subordinates
    subordinates = await employees_col().find(
        {"supervisor_id": employee_id}, {"_id": 0, "id": 1, "name": 1, "jabatan_name": 1}
    ).to_list(length=100)
    
    # Get same department colleagues
    colleagues = await employees_col().find(
        {"department": emp.get("department"), "id": {"$ne": employee_id}},
        {"_id": 0, "id": 1, "name": 1, "jabatan_name": 1}
    ).to_list(length=50)
    
    return {
        "employee_id": employee_id,
        "employee_name": emp.get("name"),
        "department": emp.get("department"),
        "supervisor": supervisor,
        "subordinates": subordinates,
        "colleagues": colleagues[:10]
    }

@router.put("/employees/{employee_id}/supervisor")
async def set_supervisor(employee_id: str, supervisor_id: str):
    """Set employee supervisor"""
    await employees_col().update_one(
        {"id": employee_id},
        {"$set": {"supervisor_id": supervisor_id, "updated_at": now_iso()}}
    )
    return {"message": "Supervisor berhasil diset"}

# ==================== EMPLOYEE DOCUMENTS ====================

@router.post("/employees/{employee_id}/documents")
async def upload_employee_document(
    employee_id: str,
    document_type: str = Form(...),
    document_name: str = Form(...),
    expiry_date: str = Form(""),
    notes: str = Form(""),
    file: UploadFile = File(...)
):
    """Upload employee document"""
    # Validate employee
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    # Save file
    ext = file.filename.split(".")[-1].lower()
    file_id = gen_id()
    filename = f"{employee_id}_{document_type}_{file_id}.{ext}"
    filepath = f"{UPLOAD_DIR}/documents/{filename}"
    
    async with aiofiles.open(filepath, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    doc = {
        "id": file_id,
        "employee_id": employee_id,
        "employee_name": emp.get("name"),
        "document_type": document_type,
        "document_name": document_name,
        "filename": filename,
        "original_filename": file.filename,
        "file_url": f"/api/hr/documents/{filename}",
        "file_size": len(content),
        "expiry_date": expiry_date,
        "notes": notes,
        "status": "active",
        "uploaded_at": now_iso()
    }
    
    await employee_docs_col().insert_one(doc)
    return {"message": "Dokumen berhasil diupload", "document": doc}

@router.get("/employees/{employee_id}/documents")
async def get_employee_documents(employee_id: str):
    """Get all documents for employee"""
    docs = await employee_docs_col().find({"employee_id": employee_id}, {"_id": 0}).to_list(length=50)
    
    # Check expiring documents
    today = get_wib_now().strftime("%Y-%m-%d")
    for d in docs:
        if d.get("expiry_date") and d["expiry_date"] <= today:
            d["is_expired"] = True
        elif d.get("expiry_date"):
            expiry = datetime.strptime(d["expiry_date"], "%Y-%m-%d")
            days_until = (expiry - get_wib_now().replace(tzinfo=None)).days
            d["days_until_expiry"] = days_until
            d["is_expiring_soon"] = days_until <= 30
    
    return {"documents": docs}

@router.get("/documents/{filename}")
async def get_document_file(filename: str):
    """Serve document file"""
    from fastapi.responses import FileResponse
    filepath = f"{UPLOAD_DIR}/documents/{filename}"
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File tidak ditemukan")
    return FileResponse(filepath)

# ==================== TRAINING MANAGEMENT ====================

@router.get("/trainings")
async def get_trainings(status: Optional[str] = None):
    """Get all trainings"""
    query = {}
    if status:
        query["status"] = status
    
    trainings = await trainings_col().find(query, {"_id": 0}).sort("start_date", -1).to_list(length=100)
    
    # Add participant count
    for t in trainings:
        count = await training_participants_col().count_documents({"training_id": t.get("id")})
        t["participant_count"] = count
    
    return {"trainings": trainings}

@router.post("/trainings")
async def create_training(training: Training):
    """Create training"""
    doc = {
        "id": gen_id(),
        **training.dict(),
        "status": "scheduled",  # scheduled, ongoing, completed, cancelled
        "created_at": now_iso()
    }
    await trainings_col().insert_one(doc)
    doc.pop("_id", None)
    return {"message": "Training berhasil dibuat", "training": doc}

@router.put("/trainings/{training_id}")
async def update_training(training_id: str, training: Training):
    """Update training"""
    await trainings_col().update_one(
        {"id": training_id},
        {"$set": {**training.dict(), "updated_at": now_iso()}}
    )
    return {"message": "Training berhasil diupdate"}

@router.post("/trainings/{training_id}/register/{employee_id}")
async def register_training(training_id: str, employee_id: str):
    """Register employee for training"""
    training = await trainings_col().find_one({"id": training_id}, {"_id": 0})
    if not training:
        raise HTTPException(status_code=404, detail="Training tidak ditemukan")
    
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    # Check if already registered
    existing = await training_participants_col().find_one({
        "training_id": training_id,
        "employee_id": employee_id
    })
    if existing:
        raise HTTPException(status_code=400, detail="Karyawan sudah terdaftar")
    
    # Check max participants
    count = await training_participants_col().count_documents({"training_id": training_id})
    if count >= training.get("max_participants", 50):
        raise HTTPException(status_code=400, detail="Training sudah penuh")
    
    participant = {
        "id": gen_id(),
        "training_id": training_id,
        "training_title": training.get("title"),
        "employee_id": employee_id,
        "employee_name": emp.get("name"),
        "department": emp.get("department"),
        "status": "registered",  # registered, attended, completed, absent
        "score": None,
        "certificate_url": "",
        "registered_at": now_iso()
    }
    
    await training_participants_col().insert_one(participant)
    participant.pop("_id", None)
    return {"message": "Berhasil mendaftar training", "participant": participant}

@router.get("/trainings/{training_id}/participants")
async def get_training_participants(training_id: str):
    """Get training participants"""
    participants = await training_participants_col().find(
        {"training_id": training_id}, {"_id": 0}
    ).to_list(length=200)
    return {"participants": participants}

@router.put("/trainings/{training_id}/participants/{employee_id}/complete")
async def complete_training(training_id: str, employee_id: str, score: Optional[float] = None):
    """Mark participant as completed"""
    await training_participants_col().update_one(
        {"training_id": training_id, "employee_id": employee_id},
        {"$set": {
            "status": "completed",
            "score": score,
            "completed_at": now_iso()
        }}
    )
    return {"message": "Training selesai"}

# ==================== ORGANIZATION STRUCTURE ====================

@router.get("/organization/structure")
async def get_org_structure():
    """Get organization structure"""
    # Get all departments with hierarchy
    depts = await departments_col().find({}, {"_id": 0}).to_list(length=100)
    
    # Get all positions
    positions = await positions_col().find({}, {"_id": 0}).to_list(length=100)
    
    # Build structure
    structure = []
    for dept in depts:
        dept_structure = {
            "id": dept.get("id"),
            "name": dept.get("name"),
            "code": dept.get("code"),
            "parent_id": dept.get("parent_id"),
            "positions": []
        }
        
        # Get employees in this department grouped by position
        employees = await employees_col().find(
            {"department": dept.get("name"), "status": "active"},
            {"_id": 0}
        ).to_list(length=500)
        
        position_groups = {}
        for emp in employees:
            pos = emp.get("jabatan_name", "Unknown")
            if pos not in position_groups:
                position_groups[pos] = []
            position_groups[pos].append({
                "id": emp.get("id"),
                "name": emp.get("name"),
                "photo_url": emp.get("photo_url", "")
            })
        
        for pos_name, emps in position_groups.items():
            dept_structure["positions"].append({
                "name": pos_name,
                "employees": emps,
                "count": len(emps)
            })
        
        dept_structure["total_employees"] = len(employees)
        structure.append(dept_structure)
    
    return {"structure": structure}

@router.get("/organization/chart")
async def get_org_chart():
    """Get organization chart data"""
    # Get top management first
    top_positions = ["Owner", "Direktur", "General Manager", "Manager"]
    
    chart_data = []
    
    for pos in top_positions:
        employees = await employees_col().find(
            {"jabatan_name": {"$regex": pos, "$options": "i"}, "status": "active"},
            {"_id": 0}
        ).to_list(length=50)
        
        for emp in employees:
            node = {
                "id": emp.get("id"),
                "name": emp.get("name"),
                "position": emp.get("jabatan_name"),
                "department": emp.get("department"),
                "photo_url": emp.get("photo_url", ""),
                "subordinates": []
            }
            
            # Get subordinates
            subs = await employees_col().find(
                {"supervisor_id": emp.get("id"), "status": "active"},
                {"_id": 0, "id": 1, "name": 1, "jabatan_name": 1, "photo_url": 1}
            ).to_list(length=50)
            
            node["subordinates"] = subs
            chart_data.append(node)
    
    return {"chart": chart_data}

# ==================== EMPLOYEE CONTRACTS ====================

@router.get("/employees/{employee_id}/contracts")
async def get_employee_contracts(employee_id: str):
    """Get employee contract history"""
    contracts = await contracts_col().find(
        {"employee_id": employee_id}, {"_id": 0}
    ).sort("start_date", -1).to_list(length=20)
    return {"contracts": contracts}

@router.post("/employees/{employee_id}/contracts")
async def create_contract(employee_id: str, contract: EmployeeContract):
    """Create new contract for employee"""
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    doc = {
        "id": gen_id(),
        **contract.dict(),
        "employee_name": emp.get("name"),
        "status": "active",
        "created_at": now_iso()
    }
    
    await contracts_col().insert_one(doc)
    
    # Update employee contract info
    await employees_col().update_one(
        {"id": employee_id},
        {"$set": {
            "contract_type": contract.contract_type,
            "contract_end_date": contract.end_date,
            "gaji_pokok": contract.salary,
            "updated_at": now_iso()
        }}
    )
    
    return {"message": "Kontrak berhasil dibuat", "contract": doc}

# ==================== WARNINGS & PROMOTIONS ====================

@router.get("/employees/{employee_id}/warnings")
async def get_employee_warnings(employee_id: str):
    """Get employee warnings"""
    warnings = await warnings_col().find(
        {"employee_id": employee_id}, {"_id": 0}
    ).sort("issued_date", -1).to_list(length=20)
    return {"warnings": warnings}

@router.post("/employees/{employee_id}/warnings")
async def create_warning(employee_id: str, warning: EmployeeWarning):
    """Issue warning to employee"""
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    doc = {
        "id": gen_id(),
        **warning.dict(),
        "employee_name": emp.get("name"),
        "status": "active",
        "created_at": now_iso()
    }
    
    await warnings_col().insert_one(doc)
    return {"message": "Peringatan berhasil dibuat", "warning": doc}

@router.get("/employees/{employee_id}/promotions")
async def get_employee_promotions(employee_id: str):
    """Get employee promotion history"""
    promotions = await promotions_col().find(
        {"employee_id": employee_id}, {"_id": 0}
    ).sort("effective_date", -1).to_list(length=20)
    return {"promotions": promotions}

@router.post("/employees/{employee_id}/promotions")
async def create_promotion(employee_id: str, promotion: EmployeePromotion):
    """Create promotion for employee"""
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    # Get position names
    from_pos = await positions_col().find_one({"id": promotion.from_position_id}, {"_id": 0})
    to_pos = await positions_col().find_one({"id": promotion.to_position_id}, {"_id": 0})
    
    doc = {
        "id": gen_id(),
        **promotion.dict(),
        "employee_name": emp.get("name"),
        "from_position_name": from_pos.get("name") if from_pos else "",
        "to_position_name": to_pos.get("name") if to_pos else "",
        "status": "pending",  # pending, approved, rejected
        "created_at": now_iso()
    }
    
    await promotions_col().insert_one(doc)
    return {"message": "Promosi berhasil diajukan", "promotion": doc}

@router.put("/promotions/{promotion_id}/approve")
async def approve_promotion(promotion_id: str, approved_by: str):
    """Approve promotion"""
    promo = await promotions_col().find_one({"id": promotion_id}, {"_id": 0})
    if not promo:
        raise HTTPException(status_code=404, detail="Promosi tidak ditemukan")
    
    # Update promotion
    await promotions_col().update_one(
        {"id": promotion_id},
        {"$set": {
            "status": "approved",
            "approved_by": approved_by,
            "approved_at": now_iso()
        }}
    )
    
    # Update employee
    await employees_col().update_one(
        {"id": promo["employee_id"]},
        {"$set": {
            "jabatan_id": promo["to_position_id"],
            "jabatan_name": promo.get("to_position_name", ""),
            "gaji_pokok": promo["to_salary"],
            "updated_at": now_iso()
        }}
    )
    
    return {"message": "Promosi berhasil disetujui"}

# ==================== MASS UPLOAD EMPLOYEES ====================

@router.get("/employees/upload-template")
async def get_employee_upload_template():
    """Get Excel template for mass employee upload"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Karyawan"
    
    # Headers
    headers = [
        'nik', 'name', 'email', 'phone', 'jabatan', 'department', 
        'branch', 'join_date', 'contract_type', 'gaji_pokok'
    ]
    
    header_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Sample data
    sample_data = [
        ['EMP-001', 'Budi Santoso', 'budi@email.com', '081234567890', 'Supervisor', 
         'Operasional', 'OCB Banjarmasin', '2024-01-15', 'tetap', 5000000],
        ['EMP-002', 'Siti Rahayu', 'siti@email.com', '081234567891', 'Kasir', 
         'Sales', 'OCB Banjarmasin', '2024-02-01', 'kontrak', 3500000],
    ]
    
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
    
    # Column widths
    col_widths = [15, 25, 25, 18, 15, 15, 20, 15, 12, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Instructions sheet
    ws_info = wb.create_sheet("Petunjuk")
    instructions = [
        ["PETUNJUK PENGISIAN DATA KARYAWAN"],
        [""],
        ["Kolom", "Keterangan", "Wajib"],
        ["nik", "Nomor Induk Karyawan (unique)", "Ya"],
        ["name", "Nama lengkap karyawan", "Ya"],
        ["email", "Email karyawan", "Tidak"],
        ["phone", "Nomor HP", "Tidak"],
        ["jabatan", "Nama jabatan", "Tidak"],
        ["department", "Nama departemen", "Tidak"],
        ["branch", "Nama cabang", "Tidak"],
        ["join_date", "Tanggal masuk (YYYY-MM-DD)", "Tidak"],
        ["contract_type", "Tipe kontrak: tetap/kontrak/magang/freelance", "Tidak"],
        ["gaji_pokok", "Gaji pokok (angka)", "Tidak"],
    ]
    
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, value in enumerate(row_data, 1):
            ws_info.cell(row=row_idx, column=col_idx, value=value)
    
    ws_info.column_dimensions['A'].width = 15
    ws_info.column_dimensions['B'].width = 45
    ws_info.column_dimensions['C'].width = 10
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={"Content-Disposition": "attachment; filename=template_karyawan.xlsx"}
    )

@router.post("/employees/mass-upload")
async def mass_upload_employees(file: UploadFile = File(...)):
    """Mass upload employees from Excel/CSV"""
    import pandas as pd
    import io
    
    content = await file.read()
    
    try:
        if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
            df = pd.read_excel(io.BytesIO(content))
        else:
            df = pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Gagal membaca file: {str(e)}")
    
    required_cols = ['nik', 'name']
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise HTTPException(status_code=400, detail=f"Kolom wajib tidak ditemukan: {', '.join(missing)}")
    
    results = {"inserted": 0, "updated": 0, "failed": 0, "errors": []}
    
    for idx, row in df.iterrows():
        try:
            nik = str(row.get('nik', '')).strip()
            if not nik:
                results["failed"] += 1
                results["errors"].append({"row": idx + 2, "error": "NIK kosong"})
                continue
            
            # Check existing
            existing = await employees_col().find_one({"nik": nik})
            
            emp_data = {
                "nik": nik,
                "name": str(row.get('name', '')).strip(),
                "email": str(row.get('email', '')).strip() if pd.notna(row.get('email')) else '',
                "phone": str(row.get('phone', '')).strip() if pd.notna(row.get('phone')) else '',
                "jabatan_name": str(row.get('jabatan', '')).strip() if pd.notna(row.get('jabatan')) else '',
                "department": str(row.get('department', '')).strip() if pd.notna(row.get('department')) else '',
                "branch_name": str(row.get('branch', '')).strip() if pd.notna(row.get('branch')) else '',
                "join_date": str(row.get('join_date', '')).strip() if pd.notna(row.get('join_date')) else '',
                "contract_type": str(row.get('contract_type', 'tetap')).strip() if pd.notna(row.get('contract_type')) else 'tetap',
                "gaji_pokok": float(row.get('gaji_pokok', 0)) if pd.notna(row.get('gaji_pokok')) else 0,
                "status": "active",
                "updated_at": now_iso()
            }
            
            if existing:
                await employees_col().update_one({"nik": nik}, {"$set": emp_data})
                results["updated"] += 1
            else:
                emp_data["id"] = gen_id()
                emp_data["created_at"] = now_iso()
                await employees_col().insert_one(emp_data)
                results["inserted"] += 1
                
        except Exception as e:
            results["failed"] += 1
            results["errors"].append({"row": idx + 2, "error": str(e)})
    
    return {
        "message": f"Upload selesai: {results['inserted']} baru, {results['updated']} update, {results['failed']} gagal",
        "results": results
    }

# ==================== HR DOCUMENT GENERATOR ====================

@router.get("/documents/generate/{doc_type}/{employee_id}")
async def generate_hr_document(doc_type: str, employee_id: str):
    """Generate HR document (SK, Kontrak, etc.)"""
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    templates = {
        "sk_pengangkatan": f"""
SURAT KEPUTUSAN PENGANGKATAN KARYAWAN
Nomor: SK/{get_wib_now().strftime('%Y%m%d')}/{emp.get('nik')}

Yang bertanda tangan di bawah ini:
Nama    : [Nama Direktur]
Jabatan : Direktur OCB GROUP

Dengan ini memutuskan mengangkat:
Nama    : {emp.get('name')}
NIK     : {emp.get('nik')}
Jabatan : {emp.get('jabatan_name')}
Departemen: {emp.get('department')}

Terhitung mulai tanggal: {emp.get('join_date')}
Status  : {emp.get('contract_type', 'Tetap').upper()}
Gaji Pokok: Rp {emp.get('gaji_pokok', 0):,.0f}

Demikian surat keputusan ini dibuat untuk dipergunakan sebagaimana mestinya.

                                    {get_wib_now().strftime('%d %B %Y')}
                                    OCB GROUP

                                    _________________
                                    Direktur
""",
        "surat_peringatan": f"""
SURAT PERINGATAN
Nomor: SP/{get_wib_now().strftime('%Y%m%d')}/{emp.get('nik')}

Kepada Yth.
{emp.get('name')}
NIK: {emp.get('nik')}
Jabatan: {emp.get('jabatan_name')}

Dengan ini kami sampaikan Surat Peringatan [TINGKAT] karena:
[ALASAN PERINGATAN]

Kami berharap Saudara dapat memperbaiki kinerja dan tidak mengulangi hal serupa.

                                    {get_wib_now().strftime('%d %B %Y')}
                                    HRD OCB GROUP

                                    _________________
                                    Manager HRD
""",
        "surat_referensi": f"""
SURAT KETERANGAN KERJA
Nomor: REF/{get_wib_now().strftime('%Y%m%d')}/{emp.get('nik')}

Yang bertanda tangan di bawah ini menerangkan bahwa:

Nama        : {emp.get('name')}
NIK         : {emp.get('nik')}
Jabatan     : {emp.get('jabatan_name')}
Departemen  : {emp.get('department')}

Adalah benar karyawan OCB GROUP sejak {emp.get('join_date')} sampai dengan sekarang
dengan status {emp.get('contract_type', 'Tetap')}.

Demikian surat keterangan ini dibuat dengan sebenarnya untuk dipergunakan sebagaimana mestinya.

                                    {get_wib_now().strftime('%d %B %Y')}
                                    OCB GROUP

                                    _________________
                                    HRD Manager
"""
    }
    
    if doc_type not in templates:
        raise HTTPException(status_code=404, detail="Template tidak ditemukan")
    
    return {
        "document_type": doc_type,
        "employee": emp.get("name"),
        "content": templates[doc_type],
        "generated_at": now_iso()
    }

# ==================== SEED DEFAULT DEPARTMENTS ====================

@router.post("/seed-departments")
async def seed_default_departments():
    """Create default departments"""
    depts = [
        {"code": "DIR", "name": "Direksi", "parent_id": "", "description": "Jajaran direksi"},
        {"code": "OPS", "name": "Operasional", "parent_id": "", "description": "Operasional toko"},
        {"code": "FIN", "name": "Keuangan", "parent_id": "", "description": "Keuangan & Akuntansi"},
        {"code": "HRD", "name": "HRD", "parent_id": "", "description": "Human Resources"},
        {"code": "IT", "name": "IT", "parent_id": "", "description": "Information Technology"},
        {"code": "MKT", "name": "Marketing", "parent_id": "", "description": "Marketing & Promosi"},
        {"code": "GUD", "name": "Gudang", "parent_id": "OPS", "description": "Pergudangan"},
        {"code": "PUR", "name": "Purchasing", "parent_id": "OPS", "description": "Pembelian"},
    ]
    
    created = 0
    for d in depts:
        existing = await departments_col().find_one({"code": d["code"]})
        if not existing:
            doc = {"id": gen_id(), **d, "is_active": True, "created_at": now_iso()}
            await departments_col().insert_one(doc)
            created += 1
    
    return {"message": f"{created} department berhasil dibuat"}
