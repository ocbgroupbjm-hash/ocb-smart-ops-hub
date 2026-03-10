# OCB TITAN AI - Payroll File Generation
# Generate payslip, rekap gaji, dll dalam format PDF, Excel, CSV, JSON

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
import uuid
import os
import json

# For PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False

# For Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

router = APIRouter(prefix="/api/payroll-files", tags=["Payroll File Generation"])

def get_db():
    from database import get_db as db_get
    return db_get()

def gen_id():
    return str(uuid.uuid4())

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def format_rupiah(amount):
    return f"Rp {amount:,.0f}".replace(",", ".")

# Collections
def employees_col():
    return get_db()['employees']

def branches_col():
    return get_db()['branches']

def payroll_slips_col():
    return get_db()['payroll_slips']

# Ensure output directory exists
OUTPUT_DIR = "/app/exports/payroll"
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ==================== CALCULATE TAKE HOME PAY ====================

def calculate_thp(emp: dict, working_days: int = 26) -> dict:
    """Calculate Take Home Pay for an employee"""
    salary_type = emp.get("salary_type", "monthly")
    
    # Base salary
    if salary_type == "daily":
        gaji_pokok = (emp.get("upah_harian", 0) or 0) * working_days
    else:
        gaji_pokok = emp.get("gaji_pokok", 0) or 0
    
    # Tunjangan
    tunjangan_jabatan = emp.get("tunjangan_jabatan", 0) or 0
    tunjangan_transport = emp.get("tunjangan_transport", 0) or 0
    tunjangan_makan = emp.get("tunjangan_makan", 0) or 0
    tunjangan_keluarga = emp.get("tunjangan_keluarga", 0) or 0
    tunjangan_lainnya = emp.get("tunjangan_lainnya", 0) or 0
    total_tunjangan = tunjangan_jabatan + tunjangan_transport + tunjangan_makan + tunjangan_keluarga + tunjangan_lainnya
    
    # Bonus
    bonus_kehadiran = emp.get("bonus_kehadiran", 0) or 0
    bonus_performance = emp.get("bonus_performance", 0) or 0
    bonus_target = emp.get("bonus_target", 0) or 0
    bonus_lainnya = emp.get("bonus_lainnya", 0) or 0
    total_bonus = bonus_kehadiran + bonus_performance + bonus_target + bonus_lainnya
    
    # Potongan
    potongan_bpjs_kes = emp.get("potongan_bpjs_kes", 0) or 0
    potongan_bpjs_tk = emp.get("potongan_bpjs_tk", 0) or 0
    potongan_pinjaman = emp.get("potongan_pinjaman", 0) or 0
    potongan_lainnya = emp.get("potongan_lainnya", 0) or 0
    total_potongan = potongan_bpjs_kes + potongan_bpjs_tk + potongan_pinjaman + potongan_lainnya
    
    # Calculate
    gross = gaji_pokok + total_tunjangan + total_bonus
    take_home_pay = gross - total_potongan
    
    return {
        "salary_type": salary_type,
        "working_days": working_days if salary_type == "daily" else None,
        "gaji_pokok": gaji_pokok,
        "tunjangan": {
            "jabatan": tunjangan_jabatan,
            "transport": tunjangan_transport,
            "makan": tunjangan_makan,
            "keluarga": tunjangan_keluarga,
            "lainnya": tunjangan_lainnya,
            "total": total_tunjangan
        },
        "bonus": {
            "kehadiran": bonus_kehadiran,
            "performance": bonus_performance,
            "target": bonus_target,
            "lainnya": bonus_lainnya,
            "total": total_bonus
        },
        "potongan": {
            "bpjs_kes": potongan_bpjs_kes,
            "bpjs_tk": potongan_bpjs_tk,
            "pinjaman": potongan_pinjaman,
            "lainnya": potongan_lainnya,
            "total": total_potongan
        },
        "gross": gross,
        "take_home_pay": take_home_pay
    }


# ==================== PAYSLIP GENERATION ====================

@router.get("/payslip/{employee_id}")
async def generate_payslip(
    employee_id: str,
    period_month: int,
    period_year: int,
    format: str = "json",  # json, pdf, excel
    working_days: int = 26
):
    """Generate payslip for a single employee using automatic payroll calculation"""
    emp = await employees_col().find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Karyawan tidak ditemukan")
    
    # Try to get automatic calculation from payroll_auto first
    try:
        from routes.payroll_auto import calculate_employee_payroll
        auto_payroll = await calculate_employee_payroll(employee_id, period_month, period_year, working_days)
        
        payslip_data = {
            "slip_id": gen_id(),
            "period": f"{period_month:02d}/{period_year}",
            "period_month": period_month,
            "period_year": period_year,
            "employee": auto_payroll["employee"],
            "attendance": auto_payroll["attendance"],
            "calculation": auto_payroll["calculation"],
            "generated_at": now_iso()
        }
    except Exception:
        # Fallback to basic calculation if auto payroll fails
        thp = calculate_thp(emp, working_days)
        payslip_data = {
            "slip_id": gen_id(),
            "period": f"{period_month:02d}/{period_year}",
            "period_month": period_month,
            "period_year": period_year,
            "employee": {
                "id": emp.get("id"),
                "nik": emp.get("nik"),
                "name": emp.get("name"),
                "jabatan": emp.get("jabatan_name"),
                "department": emp.get("department"),
                "branch": emp.get("branch_name"),
                "bank_name": emp.get("bank_name"),
                "bank_account": emp.get("bank_account"),
                "bank_holder": emp.get("bank_holder"),
                "payment_method": emp.get("payment_method", "transfer")
            },
            "attendance": None,
            "calculation": thp,
            "generated_at": now_iso()
        }
    
    if format == "json":
        return payslip_data
    
    # Get calculation data for PDF/Excel
    calc = payslip_data["calculation"]
    att = payslip_data.get("attendance") or {}
    emp_info = payslip_data["employee"]
    
    if format == "pdf":
        if not HAS_REPORTLAB:
            raise HTTPException(status_code=500, detail="PDF generation library not available")
        
        filename = f"payslip_{emp_info.get('nik')}_{period_month:02d}_{period_year}.pdf"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        # Create PDF
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        elements.append(Paragraph("SLIP GAJI KARYAWAN", styles['Title']))
        elements.append(Paragraph(f"Periode: {period_month:02d}/{period_year}", styles['Normal']))
        elements.append(Spacer(1, 20))
        
        # Employee info
        emp_info_table = [
            ["NIK:", emp_info.get("nik", "-"), "Nama:", emp_info.get("name", "-")],
            ["Jabatan:", emp_info.get("jabatan", "-"), "Cabang:", emp_info.get("branch", "-")],
            ["Bank:", emp_info.get("bank_name", "-"), "No. Rek:", emp_info.get("bank_account", "-")],
        ]
        t1 = Table(emp_info_table, colWidths=[70, 140, 70, 140])
        t1.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(t1)
        elements.append(Spacer(1, 10))
        
        # Attendance summary if available
        if att:
            att_data = [
                ["REKAP ABSENSI", "", "", ""],
                ["Hadir:", str(att.get("total_hadir", 0)), "Alpha:", str(att.get("total_alpha", 0))],
                ["Telat:", f"{att.get('total_telat_menit', 0)} menit", "Lembur:", f"{att.get('total_lembur_menit', 0)} menit"],
            ]
            t_att = Table(att_data, colWidths=[70, 140, 70, 140])
            t_att.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t_att)
            elements.append(Spacer(1, 10))
        
        # Salary breakdown
        tunjangan = calc.get("tunjangan", {})
        bonus = calc.get("bonus", {})
        potongan = calc.get("potongan", {})
        
        salary_data = [
            ["PENDAPATAN", "", "POTONGAN", ""],
            ["Gaji Dasar", format_rupiah(calc.get("gaji_dasar", 0)), "Pot. Telat", format_rupiah(potongan.get("telat", 0))],
            ["Tunj. Jabatan", format_rupiah(tunjangan.get("jabatan", 0)), "Pot. Alpha", format_rupiah(potongan.get("alpha", 0))],
            ["Tunj. Transport", format_rupiah(tunjangan.get("transport", 0)), "BPJS Kesehatan", format_rupiah(potongan.get("bpjs_kes", 0))],
            ["Tunj. Makan", format_rupiah(tunjangan.get("makan", 0)), "BPJS TK", format_rupiah(potongan.get("bpjs_tk", 0))],
            ["Tunj. Keluarga", format_rupiah(tunjangan.get("keluarga", 0)), "Pot. Pinjaman", format_rupiah(potongan.get("pinjaman", 0))],
            ["Tunj. Lainnya", format_rupiah(tunjangan.get("lainnya", 0)), "Pot. Lainnya", format_rupiah(potongan.get("lainnya", 0))],
            ["Bonus Kehadiran", format_rupiah(bonus.get("kehadiran", 0)), "", ""],
            ["Bonus Lembur", format_rupiah(bonus.get("lembur", 0)), "", ""],
            ["Bonus Penjualan", format_rupiah(bonus.get("penjualan", 0)), "", ""],
            ["Bonus Performance", format_rupiah(bonus.get("performance", 0)), "", ""],
            ["Bonus Lainnya", format_rupiah(bonus.get("lainnya", 0)), "", ""],
            ["", "", "", ""],
            ["TOTAL PENDAPATAN", format_rupiah(calc.get("gross", 0)), "TOTAL POTONGAN", format_rupiah(potongan.get("total", 0))],
        ]
        
        t2 = Table(salary_data, colWidths=[120, 90, 120, 90])
        t2.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (1, 0), colors.grey),
            ('BACKGROUND', (2, 0), (3, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t2)
        elements.append(Spacer(1, 20))
        
        # Take home pay
        thp_data = [["TAKE HOME PAY", format_rupiah(calc.get("take_home_pay", 0))]]
        t3 = Table(thp_data, colWidths=[200, 200])
        t3.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(t3)
        
        doc.build(elements)
        
        return FileResponse(filepath, filename=filename, media_type="application/pdf")
    
    elif format == "excel":
        if not HAS_OPENPYXL:
            raise HTTPException(status_code=500, detail="Excel generation library not available")
        
        filename = f"payslip_{emp_info.get('nik')}_{period_month:02d}_{period_year}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip"
        
        tunjangan = calc.get("tunjangan", {})
        bonus = calc.get("bonus", {})
        potongan = calc.get("potongan", {})
        
        # Headers
        ws['A1'] = "SLIP GAJI KARYAWAN"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Periode: {period_month:02d}/{period_year}"
        
        # Employee info
        ws['A4'] = "NIK"
        ws['B4'] = emp_info.get("nik")
        ws['C4'] = "Nama"
        ws['D4'] = emp_info.get("name")
        ws['A5'] = "Jabatan"
        ws['B5'] = emp_info.get("jabatan")
        ws['C5'] = "Cabang"
        ws['D5'] = emp_info.get("branch")
        
        # Attendance if available
        if att:
            ws['A7'] = "REKAP ABSENSI"
            ws['A7'].font = Font(bold=True)
            ws['A8'] = "Hadir"
            ws['B8'] = att.get("total_hadir", 0)
            ws['C8'] = "Alpha"
            ws['D8'] = att.get("total_alpha", 0)
            ws['A9'] = "Telat (menit)"
            ws['B9'] = att.get("total_telat_menit", 0)
            ws['C9'] = "Lembur (menit)"
            ws['D9'] = att.get("total_lembur_menit", 0)
            row = 11
        else:
            row = 7
        
        # Salary breakdown
        ws[f'A{row}'] = "PENDAPATAN"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'C{row}'] = "POTONGAN"
        ws[f'C{row}'].font = Font(bold=True)
        
        items = [
            ("Gaji Dasar", calc.get("gaji_dasar", 0), "Pot. Telat", potongan.get("telat", 0)),
            ("Tunj. Jabatan", tunjangan.get("jabatan", 0), "Pot. Alpha", potongan.get("alpha", 0)),
            ("Tunj. Transport", tunjangan.get("transport", 0), "BPJS Kesehatan", potongan.get("bpjs_kes", 0)),
            ("Tunj. Makan", tunjangan.get("makan", 0), "BPJS TK", potongan.get("bpjs_tk", 0)),
            ("Tunj. Keluarga", tunjangan.get("keluarga", 0), "Pot. Pinjaman", potongan.get("pinjaman", 0)),
            ("Tunj. Lainnya", tunjangan.get("lainnya", 0), "Pot. Lainnya", potongan.get("lainnya", 0)),
            ("Bonus Kehadiran", bonus.get("kehadiran", 0), "", ""),
            ("Bonus Lembur", bonus.get("lembur", 0), "", ""),
            ("Bonus Penjualan", bonus.get("penjualan", 0), "", ""),
            ("Bonus Performance", bonus.get("performance", 0), "", ""),
            ("Bonus Lainnya", bonus.get("lainnya", 0), "", ""),
        ]
        
        for idx, (p_name, p_val, d_name, d_val) in enumerate(items, start=row+1):
            ws[f'A{idx}'] = p_name
            ws[f'B{idx}'] = p_val
            if d_name:
                ws[f'C{idx}'] = d_name
                ws[f'D{idx}'] = d_val
        
        row = row + len(items) + 2
        ws[f'A{row}'] = "TOTAL PENDAPATAN"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = calc.get("gross", 0)
        ws[f'C{row}'] = "TOTAL POTONGAN"
        ws[f'C{row}'].font = Font(bold=True)
        ws[f'D{row}'] = potongan.get("total", 0)
        
        row += 2
        ws[f'A{row}'] = "TAKE HOME PAY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="008000")
        ws[f'B{row}'] = calc.get("take_home_pay", 0)
        ws[f'B{row}'].font = Font(bold=True, size=12, color="008000")
        
        wb.save(filepath)
        
        return FileResponse(filepath, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
    else:
        raise HTTPException(status_code=400, detail="Format tidak valid. Pilih: json, pdf, excel")


# ==================== BULK PAYROLL REPORT ====================

@router.get("/report/branch/{branch_id}")
async def generate_branch_payroll_report(
    branch_id: str,
    period_month: int,
    period_year: int,
    format: str = "json",  # json, excel, csv
    working_days: int = 26
):
    """Generate payroll report for a branch"""
    employees = await employees_col().find({
        "branch_id": branch_id,
        "status": "active"
    }, {"_id": 0}).to_list(length=1000)
    
    if not employees:
        raise HTTPException(status_code=404, detail="Tidak ada karyawan di cabang ini")
    
    branch = await branches_col().find_one({"id": branch_id}, {"_id": 0})
    branch_name = branch.get("name") if branch else "Unknown"
    
    report_data = []
    total_gross = 0
    total_potongan = 0
    total_thp = 0
    
    for emp in employees:
        thp = calculate_thp(emp, working_days)
        report_data.append({
            "nik": emp.get("nik"),
            "name": emp.get("name"),
            "jabatan": emp.get("jabatan_name"),
            "salary_type": thp["salary_type"],
            "gaji_pokok": thp["gaji_pokok"],
            "total_tunjangan": thp["tunjangan"]["total"],
            "total_bonus": thp["bonus"]["total"],
            "gross": thp["gross"],
            "total_potongan": thp["potongan"]["total"],
            "take_home_pay": thp["take_home_pay"],
            "payment_method": emp.get("payment_method", "transfer"),
            "bank": f"{emp.get('bank_name', '-')} - {emp.get('bank_account', '-')}"
        })
        total_gross += thp["gross"]
        total_potongan += thp["potongan"]["total"]
        total_thp += thp["take_home_pay"]
    
    result = {
        "report_type": "branch_payroll",
        "branch_id": branch_id,
        "branch_name": branch_name,
        "period": f"{period_month:02d}/{period_year}",
        "employee_count": len(report_data),
        "summary": {
            "total_gross": total_gross,
            "total_potongan": total_potongan,
            "total_thp": total_thp
        },
        "employees": report_data,
        "generated_at": now_iso()
    }
    
    if format == "json":
        return result
    
    elif format == "excel":
        if not HAS_OPENPYXL:
            raise HTTPException(status_code=500, detail="Excel generation library not available")
        
        filename = f"payroll_report_{branch_id[:8]}_{period_month:02d}_{period_year}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Report"
        
        # Header
        ws['A1'] = f"REKAP GAJI CABANG: {branch_name}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Periode: {period_month:02d}/{period_year}"
        
        # Table header
        headers = ["No", "NIK", "Nama", "Jabatan", "Jenis Gaji", "Gaji Pokok", "Tunjangan", "Bonus", "Gross", "Potongan", "Take Home Pay", "Metode", "Bank"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        for idx, emp in enumerate(report_data, start=5):
            ws.cell(row=idx, column=1, value=idx-4)
            ws.cell(row=idx, column=2, value=emp["nik"])
            ws.cell(row=idx, column=3, value=emp["name"])
            ws.cell(row=idx, column=4, value=emp["jabatan"])
            ws.cell(row=idx, column=5, value="Bulanan" if emp["salary_type"] == "monthly" else "Harian")
            ws.cell(row=idx, column=6, value=emp["gaji_pokok"])
            ws.cell(row=idx, column=7, value=emp["total_tunjangan"])
            ws.cell(row=idx, column=8, value=emp["total_bonus"])
            ws.cell(row=idx, column=9, value=emp["gross"])
            ws.cell(row=idx, column=10, value=emp["total_potongan"])
            ws.cell(row=idx, column=11, value=emp["take_home_pay"])
            ws.cell(row=idx, column=12, value=emp["payment_method"])
            ws.cell(row=idx, column=13, value=emp["bank"])
        
        # Total
        total_row = len(report_data) + 5
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=total_gross).font = Font(bold=True)
        ws.cell(row=total_row, column=10, value=total_potongan).font = Font(bold=True)
        ws.cell(row=total_row, column=11, value=total_thp).font = Font(bold=True, color="008000")
        
        wb.save(filepath)
        return FileResponse(filepath, filename=filename)
    
    elif format == "csv":
        import csv
        import io
        
        filename = f"payroll_report_{branch_id[:8]}_{period_month:02d}_{period_year}.csv"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["NIK", "Nama", "Jabatan", "Jenis Gaji", "Gaji Pokok", "Tunjangan", "Bonus", "Gross", "Potongan", "Take Home Pay", "Metode", "Bank"])
            for emp in report_data:
                writer.writerow([
                    emp["nik"], emp["name"], emp["jabatan"], emp["salary_type"],
                    emp["gaji_pokok"], emp["total_tunjangan"], emp["total_bonus"],
                    emp["gross"], emp["total_potongan"], emp["take_home_pay"],
                    emp["payment_method"], emp["bank"]
                ])
        
        return FileResponse(filepath, filename=filename, media_type="text/csv")
    
    else:
        raise HTTPException(status_code=400, detail="Format tidak valid")


# ==================== COMPANY WIDE PAYROLL REPORT ====================

@router.get("/report/company")
async def generate_company_payroll_report(
    period_month: int,
    period_year: int,
    format: str = "json",
    working_days: int = 26
):
    """Generate company-wide payroll report"""
    employees = await employees_col().find({"status": "active"}, {"_id": 0}).to_list(length=5000)
    branches = await branches_col().find({}, {"_id": 0}).to_list(length=100)
    branch_map = {b.get("id"): b.get("name") for b in branches}
    
    # Group by branch
    branch_data = {}
    
    for emp in employees:
        bid = emp.get("branch_id", "unknown")
        bname = branch_map.get(bid, emp.get("branch_name", "Unknown"))
        
        if bid not in branch_data:
            branch_data[bid] = {
                "branch_id": bid,
                "branch_name": bname,
                "employee_count": 0,
                "monthly_employees": 0,
                "daily_employees": 0,
                "total_gross": 0,
                "total_potongan": 0,
                "total_thp": 0
            }
        
        thp = calculate_thp(emp, working_days)
        
        branch_data[bid]["employee_count"] += 1
        if emp.get("salary_type") == "daily":
            branch_data[bid]["daily_employees"] += 1
        else:
            branch_data[bid]["monthly_employees"] += 1
        branch_data[bid]["total_gross"] += thp["gross"]
        branch_data[bid]["total_potongan"] += thp["potongan"]["total"]
        branch_data[bid]["total_thp"] += thp["take_home_pay"]
    
    branch_list = list(branch_data.values())
    branch_list.sort(key=lambda x: x["total_thp"], reverse=True)
    
    grand_total = {
        "employee_count": sum(b["employee_count"] for b in branch_list),
        "monthly_employees": sum(b["monthly_employees"] for b in branch_list),
        "daily_employees": sum(b["daily_employees"] for b in branch_list),
        "total_gross": sum(b["total_gross"] for b in branch_list),
        "total_potongan": sum(b["total_potongan"] for b in branch_list),
        "total_thp": sum(b["total_thp"] for b in branch_list)
    }
    
    result = {
        "report_type": "company_payroll",
        "period": f"{period_month:02d}/{period_year}",
        "branch_count": len(branch_list),
        "grand_total": grand_total,
        "branches": branch_list,
        "generated_at": now_iso()
    }
    
    if format == "json":
        return result
    
    elif format == "excel":
        if not HAS_OPENPYXL:
            raise HTTPException(status_code=500, detail="Excel generation library not available")
        
        filename = f"payroll_company_{period_month:02d}_{period_year}.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Company Payroll"
        
        # Header
        ws['A1'] = "REKAP GAJI PERUSAHAAN"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Periode: {period_month:02d}/{period_year}"
        
        # Table header
        headers = ["No", "Cabang", "Jumlah Karyawan", "Bulanan", "Harian", "Total Gross", "Total Potongan", "Total THP"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        for idx, br in enumerate(branch_list, start=5):
            ws.cell(row=idx, column=1, value=idx-4)
            ws.cell(row=idx, column=2, value=br["branch_name"])
            ws.cell(row=idx, column=3, value=br["employee_count"])
            ws.cell(row=idx, column=4, value=br["monthly_employees"])
            ws.cell(row=idx, column=5, value=br["daily_employees"])
            ws.cell(row=idx, column=6, value=br["total_gross"])
            ws.cell(row=idx, column=7, value=br["total_potongan"])
            ws.cell(row=idx, column=8, value=br["total_thp"])
        
        # Grand Total
        total_row = len(branch_list) + 5
        ws.cell(row=total_row, column=1, value="GRAND TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=grand_total["employee_count"]).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=grand_total["monthly_employees"]).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=grand_total["daily_employees"]).font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=grand_total["total_gross"]).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=grand_total["total_potongan"]).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=grand_total["total_thp"]).font = Font(bold=True, color="008000")
        
        wb.save(filepath)
        return FileResponse(filepath, filename=filename)
    
    else:
        raise HTTPException(status_code=400, detail="Format tidak valid")


# ==================== HR DASHBOARD SUMMARY ====================

@router.get("/dashboard-summary")
async def get_payroll_dashboard_summary():
    """Get payroll summary for HR dashboard"""
    employees = await employees_col().find({"status": "active"}, {"_id": 0}).to_list(length=5000)
    
    total_employees = len(employees)
    monthly_employees = len([e for e in employees if e.get("salary_type", "monthly") == "monthly"])
    daily_employees = len([e for e in employees if e.get("salary_type") == "daily"])
    
    total_gross = 0
    total_potongan = 0
    total_thp = 0
    
    for emp in employees:
        thp = calculate_thp(emp)
        total_gross += thp["gross"]
        total_potongan += thp["potongan"]["total"]
        total_thp += thp["take_home_pay"]
    
    return {
        "total_employees": total_employees,
        "monthly_employees": monthly_employees,
        "daily_employees": daily_employees,
        "estimated_monthly_payroll": {
            "total_gross": total_gross,
            "total_potongan": total_potongan,
            "total_take_home_pay": total_thp
        },
        "generated_at": now_iso()
    }
