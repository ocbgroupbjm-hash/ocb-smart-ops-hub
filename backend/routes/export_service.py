# OCB TITAN ERP - Export Service
# SSOT-based export engine - Data harus diambil dari source of truth
# tenant-aware, branch-aware, date-range filter

from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from database import get_db
from utils.auth import get_current_user

router = APIRouter(prefix="/export", tags=["Export"])

# Styles for Excel
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
CURRENCY_FORMAT = '#,##0'
DATE_FORMAT = 'YYYY-MM-DD'
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def create_workbook(title: str, tenant_name: str = "", date_range: str = ""):
    """Create a new workbook with standard header"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel limit
    
    # Title row
    ws['A1'] = f"OCB TITAN ERP - {title}"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    
    # Metadata
    ws['A2'] = f"Tenant: {tenant_name}" if tenant_name else ""
    ws['A3'] = f"Periode: {date_range}" if date_range else f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    return wb, ws


def style_header_row(ws, row_num: int, num_cols: int):
    """Apply header styling to a row"""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def auto_column_width(ws):
    """Auto-adjust column widths"""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 10)


# ==================== SALES EXPORT ====================

@router.get("/sales")
async def export_sales(
    date_from: str = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: str = Query(None, description="End date (YYYY-MM-DD)"),
    branch_id: str = Query(None, description="Branch ID filter"),
    status: str = Query(None, description="Status filter"),
    user: dict = Depends(get_current_user)
):
    """Export Sales Invoices to Excel - SSOT: sales_invoices collection"""
    db = get_db()
    
    # Build query
    query = {}
    
    # Branch filter (tenant-aware)
    if user.get("role") not in ["owner", "admin"]:
        if user.get("branch_id"):
            query["branch_id"] = user["branch_id"]
    elif branch_id:
        query["branch_id"] = branch_id
    
    # Date filter
    if date_from or date_to:
        query["invoice_date"] = {}
        if date_from:
            query["invoice_date"]["$gte"] = date_from
        if date_to:
            query["invoice_date"]["$lte"] = date_to + "T23:59:59"
    
    if status:
        query["status"] = status
    
    # Fetch data from SSOT
    sales = await db["sales_invoices"].find(query, {"_id": 0}).sort("invoice_date", -1).to_list(10000)
    
    if not sales:
        raise HTTPException(status_code=404, detail="Tidak ada data penjualan untuk diekspor")
    
    # Get tenant info
    tenant = await db["tenants"].find_one({}, {"_id": 0, "name": 1})
    tenant_name = tenant.get("name", "") if tenant else ""
    
    # Create workbook
    date_range = f"{date_from or 'awal'} - {date_to or 'sekarang'}"
    wb, ws = create_workbook("Data Penjualan", tenant_name, date_range)
    
    # Headers (row 5)
    headers = ["No", "Tanggal", "No. Invoice", "Customer", "Cabang", "Items", "Subtotal", "Diskon", "PPN", "Total", "Metode", "Status"]
    ws.append([])  # Row 4 spacer
    ws.append(headers)
    style_header_row(ws, 5, len(headers))
    
    # Data rows
    for idx, sale in enumerate(sales, 1):
        items_count = len(sale.get("items", []))
        row = [
            idx,
            sale.get("invoice_date", "")[:10],
            sale.get("invoice_number", ""),
            sale.get("customer_name", "-"),
            sale.get("branch_name", "-"),
            items_count,
            sale.get("subtotal", 0),
            sale.get("discount_amount", 0),
            sale.get("tax_amount", 0),
            sale.get("total", 0),
            sale.get("payment_method", "-"),
            sale.get("status", "-")
        ]
        ws.append(row)
        
        # Format currency columns
        for col in [7, 8, 9, 10]:  # Subtotal, Diskon, PPN, Total
            ws.cell(row=5 + idx, column=col).number_format = CURRENCY_FORMAT
    
    # Summary row
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=7, value=sum(s.get("subtotal", 0) for s in sales))
    ws.cell(row=summary_row, column=8, value=sum(s.get("discount_amount", 0) for s in sales))
    ws.cell(row=summary_row, column=9, value=sum(s.get("tax_amount", 0) for s in sales))
    ws.cell(row=summary_row, column=10, value=sum(s.get("total", 0) for s in sales))
    
    for col in [7, 8, 9, 10]:
        ws.cell(row=summary_row, column=col).number_format = CURRENCY_FORMAT
        ws.cell(row=summary_row, column=col).font = Font(bold=True)
    
    auto_column_width(ws)
    
    # Return as streaming response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"sales_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== PURCHASE EXPORT ====================

@router.get("/purchase")
async def export_purchase(
    date_from: str = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: str = Query(None, description="End date (YYYY-MM-DD)"),
    branch_id: str = Query(None, description="Branch ID filter"),
    status: str = Query(None, description="Status filter"),
    user: dict = Depends(get_current_user)
):
    """Export Purchase Orders to Excel - SSOT: purchase_orders collection"""
    db = get_db()
    
    query = {}
    
    if user.get("role") not in ["owner", "admin"]:
        if user.get("branch_id"):
            query["branch_id"] = user["branch_id"]
    elif branch_id:
        query["branch_id"] = branch_id
    
    if date_from or date_to:
        query["order_date"] = {}
        if date_from:
            query["order_date"]["$gte"] = date_from
        if date_to:
            query["order_date"]["$lte"] = date_to + "T23:59:59"
    
    if status:
        query["status"] = status
    
    purchases = await db["purchase_orders"].find(query, {"_id": 0}).sort("order_date", -1).to_list(10000)
    
    if not purchases:
        raise HTTPException(status_code=404, detail="Tidak ada data pembelian untuk diekspor")
    
    tenant = await db["tenants"].find_one({}, {"_id": 0, "name": 1})
    tenant_name = tenant.get("name", "") if tenant else ""
    
    date_range = f"{date_from or 'awal'} - {date_to or 'sekarang'}"
    wb, ws = create_workbook("Data Pembelian", tenant_name, date_range)
    
    headers = ["No", "Tanggal", "No. PO", "Supplier", "Cabang", "Items", "Subtotal", "Diskon", "PPN", "Total", "Kredit", "Status"]
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 5, len(headers))
    
    for idx, po in enumerate(purchases, 1):
        items_count = len(po.get("items", []))
        row = [
            idx,
            po.get("order_date", "")[:10] if po.get("order_date") else "",
            po.get("po_number", ""),
            po.get("supplier_name", "-"),
            po.get("branch_name", "-") if po.get("branch_name") else "-",
            items_count,
            po.get("subtotal", 0),
            po.get("discount_amount", 0),
            po.get("tax_amount", 0),
            po.get("total", 0),
            "Ya" if po.get("is_credit") else "Tidak",
            po.get("status", "-")
        ]
        ws.append(row)
        
        for col in [7, 8, 9, 10]:
            ws.cell(row=5 + idx, column=col).number_format = CURRENCY_FORMAT
    
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=7, value=sum(p.get("subtotal", 0) for p in purchases))
    ws.cell(row=summary_row, column=10, value=sum(p.get("total", 0) for p in purchases))
    
    for col in [7, 10]:
        ws.cell(row=summary_row, column=col).number_format = CURRENCY_FORMAT
        ws.cell(row=summary_row, column=col).font = Font(bold=True)
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"purchase_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== LEDGER EXPORT ====================

@router.get("/ledger")
async def export_ledger(
    date_from: str = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: str = Query(None, description="End date (YYYY-MM-DD)"),
    account_code: str = Query(None, description="Account code filter"),
    branch_id: str = Query(None, description="Branch ID filter"),
    user: dict = Depends(get_current_user)
):
    """Export General Ledger to Excel - SSOT: journal_entries collection"""
    db = get_db()
    
    query = {"status": "posted"}
    
    if user.get("role") not in ["owner", "admin"]:
        if user.get("branch_id"):
            query["branch_id"] = user["branch_id"]
    elif branch_id:
        query["branch_id"] = branch_id
    
    if date_from or date_to:
        query["journal_date"] = {}
        if date_from:
            query["journal_date"]["$gte"] = date_from
        if date_to:
            query["journal_date"]["$lte"] = date_to + "T23:59:59"
    
    journals = await db["journal_entries"].find(query, {"_id": 0}).sort("journal_date", 1).to_list(50000)
    
    if not journals:
        raise HTTPException(status_code=404, detail="Tidak ada data jurnal untuk diekspor")
    
    tenant = await db["tenants"].find_one({}, {"_id": 0, "name": 1})
    tenant_name = tenant.get("name", "") if tenant else ""
    
    date_range = f"{date_from or 'awal'} - {date_to or 'sekarang'}"
    wb, ws = create_workbook("Buku Besar (General Ledger)", tenant_name, date_range)
    
    headers = ["Tanggal", "No. Jurnal", "Kode Akun", "Nama Akun", "Deskripsi", "Debit", "Credit", "Referensi"]
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 5, len(headers))
    
    total_debit = 0
    total_credit = 0
    row_num = 5
    
    for journal in journals:
        entries = journal.get("entries", [])
        
        # Filter by account_code if specified
        if account_code:
            entries = [e for e in entries if e.get("account_code", "").startswith(account_code)]
        
        for entry in entries:
            row_num += 1
            debit = entry.get("debit", 0)
            credit = entry.get("credit", 0)
            total_debit += debit
            total_credit += credit
            
            row = [
                journal.get("journal_date", "")[:10],
                journal.get("journal_number", ""),
                entry.get("account_code", ""),
                entry.get("account_name", ""),
                entry.get("description", journal.get("description", "")),
                debit if debit > 0 else "",
                credit if credit > 0 else "",
                journal.get("reference_number", "")
            ]
            ws.append(row)
            
            if debit > 0:
                ws.cell(row=row_num, column=6).number_format = CURRENCY_FORMAT
            if credit > 0:
                ws.cell(row=row_num, column=7).number_format = CURRENCY_FORMAT
    
    # Summary
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=6, value=total_debit)
    ws.cell(row=summary_row, column=7, value=total_credit)
    
    for col in [6, 7]:
        ws.cell(row=summary_row, column=col).number_format = CURRENCY_FORMAT
        ws.cell(row=summary_row, column=col).font = Font(bold=True)
    
    # Balance check
    ws.append([])
    balance_row = ws.max_row + 1
    ws.cell(row=balance_row, column=1, value="SELISIH")
    ws.cell(row=balance_row, column=6, value=total_debit - total_credit)
    ws.cell(row=balance_row, column=6).number_format = CURRENCY_FORMAT
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ledger_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== TRIAL BALANCE EXPORT ====================

@router.get("/trial-balance")
async def export_trial_balance(
    as_of_date: str = Query(None, description="As of date (YYYY-MM-DD)"),
    branch_id: str = Query(None, description="Branch ID filter"),
    user: dict = Depends(get_current_user)
):
    """Export Trial Balance to Excel - SSOT: aggregated from journal_entries"""
    db = get_db()
    
    query = {"status": "posted"}
    
    if user.get("role") not in ["owner", "admin"]:
        if user.get("branch_id"):
            query["branch_id"] = user["branch_id"]
    elif branch_id:
        query["branch_id"] = branch_id
    
    if as_of_date:
        query["journal_date"] = {"$lte": as_of_date + "T23:59:59"}
    
    journals = await db["journal_entries"].find(query, {"_id": 0}).to_list(100000)
    
    # Aggregate by account
    account_balances = {}
    for journal in journals:
        for entry in journal.get("entries", []):
            code = entry.get("account_code", "")
            name = entry.get("account_name", "")
            if not code:
                continue
            
            if code not in account_balances:
                account_balances[code] = {"name": name, "debit": 0, "credit": 0}
            
            account_balances[code]["debit"] += entry.get("debit", 0)
            account_balances[code]["credit"] += entry.get("credit", 0)
    
    if not account_balances:
        raise HTTPException(status_code=404, detail="Tidak ada data untuk trial balance")
    
    tenant = await db["tenants"].find_one({}, {"_id": 0, "name": 1})
    tenant_name = tenant.get("name", "") if tenant else ""
    
    period = f"s/d {as_of_date}" if as_of_date else f"s/d {datetime.now().strftime('%Y-%m-%d')}"
    wb, ws = create_workbook("Neraca Saldo (Trial Balance)", tenant_name, period)
    
    headers = ["Kode Akun", "Nama Akun", "Debit", "Credit"]
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 5, len(headers))
    
    total_debit = 0
    total_credit = 0
    
    for code in sorted(account_balances.keys()):
        data = account_balances[code]
        debit = data["debit"]
        credit = data["credit"]
        
        # Skip zero balances
        if debit == 0 and credit == 0:
            continue
        
        total_debit += debit
        total_credit += credit
        
        row = [code, data["name"], debit, credit]
        ws.append(row)
        
        row_num = ws.max_row
        ws.cell(row=row_num, column=3).number_format = CURRENCY_FORMAT
        ws.cell(row=row_num, column=4).number_format = CURRENCY_FORMAT
    
    # Summary
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL")
    ws.cell(row=summary_row, column=3, value=total_debit)
    ws.cell(row=summary_row, column=4, value=total_credit)
    
    for col in [3, 4]:
        ws.cell(row=summary_row, column=col).number_format = CURRENCY_FORMAT
        ws.cell(row=summary_row, column=col).font = Font(bold=True)
    
    # Balance status
    ws.append([])
    status_row = ws.max_row + 1
    is_balanced = abs(total_debit - total_credit) < 0.01
    ws.cell(row=status_row, column=1, value="STATUS")
    ws.cell(row=status_row, column=2, value="SEIMBANG" if is_balanced else f"TIDAK SEIMBANG (Selisih: {total_debit - total_credit:,.2f})")
    ws.cell(row=status_row, column=2).font = Font(bold=True, color="00FF00" if is_balanced else "FF0000")
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"trial_balance_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== INVENTORY EXPORT ====================

@router.get("/inventory")
async def export_inventory(
    branch_id: str = Query(None, description="Branch ID filter"),
    category_id: str = Query(None, description="Category ID filter"),
    user: dict = Depends(get_current_user)
):
    """Export Inventory to Excel - SSOT: stock_movements collection (aggregated)"""
    db = get_db()
    
    # Get branches to export
    branch_query = {}
    if user.get("role") not in ["owner", "admin"]:
        if user.get("branch_id"):
            branch_query["id"] = user["branch_id"]
    elif branch_id:
        branch_query["id"] = branch_id
    
    branches = await db["branches"].find(branch_query, {"_id": 0}).to_list(100)
    
    if not branches:
        raise HTTPException(status_code=404, detail="Tidak ada cabang untuk diekspor")
    
    # Get products
    product_query = {"is_active": True}
    if category_id:
        product_query["category_id"] = category_id
    
    products = await db["products"].find(product_query, {"_id": 0}).to_list(10000)
    product_map = {p["id"]: p for p in products}
    
    tenant = await db["tenants"].find_one({}, {"_id": 0, "name": 1})
    tenant_name = tenant.get("name", "") if tenant else ""
    
    wb, ws = create_workbook("Inventory (Stok Barang)", tenant_name, f"Per {datetime.now().strftime('%Y-%m-%d')}")
    
    headers = ["No", "Kode Produk", "Nama Produk", "Kategori", "Cabang", "Stok (SSOT)", "Harga Beli", "Harga Jual", "Nilai Stok"]
    ws.append([])
    ws.append(headers)
    style_header_row(ws, 5, len(headers))
    
    row_idx = 0
    total_stock_value = 0
    
    for branch in branches:
        branch_id = branch["id"]
        branch_name = branch.get("name", "")
        
        # Calculate stock from SSOT (stock_movements)
        pipeline = [
            {"$match": {"branch_id": branch_id}},
            {"$group": {"_id": "$product_id", "total_qty": {"$sum": "$quantity"}}}
        ]
        stock_data = await db["stock_movements"].aggregate(pipeline).to_list(10000)
        stock_map = {s["_id"]: s["total_qty"] for s in stock_data}
        
        for product_id, qty in stock_map.items():
            if qty == 0:
                continue
            
            product = product_map.get(product_id, {})
            if not product:
                continue
            
            row_idx += 1
            cost_price = product.get("cost_price", 0)
            stock_value = qty * cost_price
            total_stock_value += stock_value
            
            row = [
                row_idx,
                product.get("code", ""),
                product.get("name", ""),
                product.get("category_name", "-"),
                branch_name,
                qty,
                cost_price,
                product.get("selling_price", 0),
                stock_value
            ]
            ws.append(row)
            
            row_num = ws.max_row
            for col in [7, 8, 9]:
                ws.cell(row=row_num, column=col).number_format = CURRENCY_FORMAT
    
    # Summary
    ws.append([])
    summary_row = ws.max_row + 1
    ws.cell(row=summary_row, column=1, value="TOTAL NILAI STOK")
    ws.cell(row=summary_row, column=9, value=total_stock_value)
    ws.cell(row=summary_row, column=9).number_format = CURRENCY_FORMAT
    ws.cell(row=summary_row, column=9).font = Font(bold=True)
    
    auto_column_width(ws)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"inventory_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
