# OCB TITAN - Reports API
from fastapi import APIRouter, HTTPException, Depends
from database import (
    transactions, products, product_stocks, customers, branches,
    stock_movements, expenses, suppliers, purchase_orders
)
from utils.auth import get_current_user
from datetime import datetime, timezone, timedelta

router = APIRouter(prefix="/reports", tags=["Reports"])

@router.get("/sales")
async def sales_report(
    date_from: str,
    date_to: str,
    branch_id: str = "",
    group_by: str = "day",  # day, week, month, branch, product, category
    user: dict = Depends(get_current_user)
):
    """Comprehensive sales report"""
    match_query = {
        "status": "completed",
        "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
    }
    
    if branch_id:
        match_query["branch_id"] = branch_id
    
    if group_by == "day":
        pipeline = [
            {"$match": match_query},
            {
                "$group": {
                    "_id": {"$substr": ["$created_at", 0, 10]},
                    "gross_sales": {"$sum": "$subtotal"},
                    "discounts": {"$sum": "$discount_amount"},
                    "net_sales": {"$sum": "$total"},
                    "cost": {"$sum": "$total_cost"},
                    "profit": {"$sum": "$profit"},
                    "transactions": {"$sum": 1}
                }
            },
            {"$sort": {"_id": 1}}
        ]
    elif group_by == "branch":
        pipeline = [
            {"$match": match_query},
            {
                "$group": {
                    "_id": "$branch_id",
                    "gross_sales": {"$sum": "$subtotal"},
                    "discounts": {"$sum": "$discount_amount"},
                    "net_sales": {"$sum": "$total"},
                    "cost": {"$sum": "$total_cost"},
                    "profit": {"$sum": "$profit"},
                    "transactions": {"$sum": 1}
                }
            },
            {"$sort": {"net_sales": -1}}
        ]
    elif group_by == "product":
        pipeline = [
            {"$match": match_query},
            {"$unwind": "$items"},
            {
                "$group": {
                    "_id": "$items.product_id",
                    "product_name": {"$first": "$items.product_name"},
                    "product_code": {"$first": "$items.product_code"},
                    "quantity": {"$sum": "$items.quantity"},
                    "revenue": {"$sum": "$items.total"},
                    "cost": {"$sum": "$items.cost_price"}
                }
            },
            {"$addFields": {"profit": {"$subtract": ["$revenue", "$cost"]}}},
            {"$sort": {"revenue": -1}},
            {"$limit": 100}
        ]
    else:  # Default to day
        pipeline = [
            {"$match": match_query},
            {
                "$group": {
                    "_id": {"$substr": ["$created_at", 0, 10]},
                    "net_sales": {"$sum": "$total"},
                    "profit": {"$sum": "$profit"},
                    "transactions": {"$sum": 1}
                }
            },
            {"$sort": {"_id": 1}}
        ]
    
    result = await transactions.aggregate(pipeline).to_list(1000)
    
    # Enrich branch names if grouped by branch
    if group_by == "branch":
        for item in result:
            branch = await branches.find_one({"id": item["_id"]}, {"_id": 0, "name": 1, "code": 1})
            item["branch_name"] = branch.get("name", "Unknown") if branch else "Unknown"
            item["branch_code"] = branch.get("code", "") if branch else ""
    
    # Calculate totals
    totals = {
        "net_sales": sum(r.get("net_sales", 0) for r in result),
        "profit": sum(r.get("profit", 0) for r in result),
        "transactions": sum(r.get("transactions", 0) for r in result)
    }
    
    return {
        "period": {"from": date_from, "to": date_to},
        "group_by": group_by,
        "data": result,
        "totals": totals
    }

@router.get("/inventory")
async def inventory_report(
    branch_id: str = "",
    category_id: str = "",
    include_zero_stock: bool = False,
    user: dict = Depends(get_current_user)
):
    """Inventory valuation report"""
    bid = branch_id or user.get("branch_id")
    
    pipeline = [
        {"$match": {"branch_id": bid} if bid else {}},
        {
            "$lookup": {
                "from": "products",
                "localField": "product_id",
                "foreignField": "id",
                "as": "product"
            }
        },
        {"$unwind": "$product"},
        {"$match": {"product.is_active": True}}
    ]
    
    if category_id:
        pipeline.append({"$match": {"product.category_id": category_id}})
    
    if not include_zero_stock:
        pipeline.append({"$match": {"quantity": {"$gt": 0}}})
    
    pipeline.extend([
        {
            "$project": {
                "_id": 0,
                "product_id": 1,
                "branch_id": 1,
                "product_code": "$product.code",
                "product_name": "$product.name",
                "category_id": "$product.category_id",
                "quantity": 1,
                "cost_price": "$product.cost_price",
                "selling_price": "$product.selling_price",
                "stock_value": {"$multiply": ["$quantity", "$product.cost_price"]},
                "retail_value": {"$multiply": ["$quantity", "$product.selling_price"]},
                "min_stock": "$product.min_stock",
                "is_low_stock": {"$lte": ["$quantity", "$product.min_stock"]}
            }
        },
        {"$sort": {"product_name": 1}}
    ])
    
    result = await product_stocks.aggregate(pipeline).to_list(10000)
    
    # Calculate totals - handle None values
    total_stock_value = sum((r.get("stock_value") or 0) for r in result)
    total_retail_value = sum((r.get("retail_value") or 0) for r in result)
    total_items = sum((r.get("quantity") or 0) for r in result)
    low_stock_count = sum(1 for r in result if r.get("is_low_stock", False))
    
    return {
        "branch_id": bid,
        "data": result,
        "summary": {
            "total_products": len(result),
            "total_items": total_items,
            "total_stock_value": total_stock_value,
            "total_retail_value": total_retail_value,
            "potential_profit": total_retail_value - total_stock_value,
            "low_stock_count": low_stock_count
        }
    }

@router.get("/stock-movements")
async def stock_movement_report(
    date_from: str,
    date_to: str,
    branch_id: str = "",
    movement_type: str = "",
    product_id: str = "",
    user: dict = Depends(get_current_user)
):
    """Stock movement history report"""
    query = {
        "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
    }
    
    if branch_id:
        query["branch_id"] = branch_id
    
    if movement_type:
        query["movement_type"] = movement_type
    
    if product_id:
        query["product_id"] = product_id
    
    items = await stock_movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    
    # Enrich with product names
    for item in items:
        product = await products.find_one({"id": item["product_id"]}, {"_id": 0, "name": 1, "code": 1})
        item["product_name"] = product.get("name", "") if product else "Unknown"
        item["product_code"] = product.get("code", "") if product else ""
    
    # Summary by type
    summary_pipeline = [
        {"$match": query},
        {
            "$group": {
                "_id": "$movement_type",
                "count": {"$sum": 1},
                "quantity": {"$sum": "$quantity"}
            }
        }
    ]
    
    summary = await stock_movements.aggregate(summary_pipeline).to_list(20)
    
    return {
        "period": {"from": date_from, "to": date_to},
        "data": items,
        "summary": {item["_id"]: {"count": item["count"], "quantity": item["quantity"]} for item in summary}
    }

@router.get("/customer-analysis")
async def customer_analysis_report(
    date_from: str = "",
    date_to: str = "",
    user: dict = Depends(get_current_user)
):
    """Customer analysis report"""
    # If no dates, use last 30 days
    if not date_to:
        date_to = datetime.now(timezone.utc).isoformat()[:10]
    if not date_from:
        date_from = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()[:10]
    
    # Customer segments
    segment_pipeline = [
        {"$match": {"is_active": True}},
        {
            "$group": {
                "_id": "$segment",
                "count": {"$sum": 1},
                "total_spent": {"$sum": "$total_spent"},
                "avg_spent": {"$avg": "$total_spent"}
            }
        }
    ]
    
    segments = await customers.aggregate(segment_pipeline).to_list(10)
    
    # Top customers by spending in period
    top_customers_pipeline = [
        {
            "$match": {
                "status": "completed",
                "customer_id": {"$ne": None},
                "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
            }
        },
        {
            "$group": {
                "_id": "$customer_id",
                "customer_name": {"$first": "$customer_name"},
                "total_spent": {"$sum": "$total"},
                "transactions": {"$sum": 1},
                "avg_transaction": {"$avg": "$total"}
            }
        },
        {"$sort": {"total_spent": -1}},
        {"$limit": 20}
    ]
    
    top_customers = await transactions.aggregate(top_customers_pipeline).to_list(20)
    
    # New customers in period
    new_customers_pipeline = [
        {
            "$match": {
                "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
            }
        },
        {"$count": "count"}
    ]
    
    new_result = await customers.aggregate(new_customers_pipeline).to_list(1)
    new_customers = new_result[0]["count"] if new_result else 0
    
    # Total customers
    total_customers = await customers.count_documents({"is_active": True})
    
    return {
        "period": {"from": date_from, "to": date_to},
        "segments": segments,
        "top_customers": top_customers,
        "new_customers": new_customers,
        "total_customers": total_customers
    }

@router.get("/product-performance")
async def product_performance_report(
    date_from: str,
    date_to: str,
    branch_id: str = "",
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    """Product performance analysis"""
    match_query = {
        "status": "completed",
        "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
    }
    
    if branch_id:
        match_query["branch_id"] = branch_id
    
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {
            "$group": {
                "_id": "$items.product_id",
                "product_name": {"$first": "$items.product_name"},
                "product_code": {"$first": "$items.product_code"},
                "quantity_sold": {"$sum": "$items.quantity"},
                "gross_revenue": {"$sum": "$items.subtotal"},
                "discounts": {"$sum": "$items.discount_amount"},
                "net_revenue": {"$sum": "$items.total"},
                "cost": {"$sum": "$items.cost_price"},
                "transactions": {"$sum": 1}
            }
        },
        {
            "$addFields": {
                "profit": {"$subtract": ["$net_revenue", "$cost"]},
                "avg_price": {"$divide": ["$net_revenue", "$quantity_sold"]},
                "profit_margin": {
                    "$multiply": [
                        {"$divide": [{"$subtract": ["$net_revenue", "$cost"]}, {"$max": ["$net_revenue", 1]}]},
                        100
                    ]
                }
            }
        },
        {"$sort": {"net_revenue": -1}},
        {"$limit": limit}
    ]
    
    result = await transactions.aggregate(pipeline).to_list(limit)
    
    # Totals
    totals = {
        "total_revenue": sum(r.get("net_revenue", 0) for r in result),
        "total_profit": sum(r.get("profit", 0) for r in result),
        "total_quantity": sum(r.get("quantity_sold", 0) for r in result)
    }
    
    return {
        "period": {"from": date_from, "to": date_to},
        "data": result,
        "totals": totals
    }

@router.get("/branch-comparison")
async def branch_comparison_report(
    date_from: str,
    date_to: str,
    user: dict = Depends(get_current_user)
):
    """Compare performance across branches"""
    pipeline = [
        {
            "$match": {
                "status": "completed",
                "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
            }
        },
        {
            "$group": {
                "_id": "$branch_id",
                "total_sales": {"$sum": "$total"},
                "total_profit": {"$sum": "$profit"},
                "total_cost": {"$sum": "$total_cost"},
                "transactions": {"$sum": 1},
                "avg_transaction": {"$avg": "$total"}
            }
        },
        {
            "$addFields": {
                "profit_margin": {
                    "$multiply": [
                        {"$divide": ["$total_profit", {"$max": ["$total_sales", 1]}]},
                        100
                    ]
                }
            }
        },
        {"$sort": {"total_sales": -1}}
    ]
    
    result = await transactions.aggregate(pipeline).to_list(500)
    
    # Enrich with branch info
    for item in result:
        branch = await branches.find_one({"id": item["_id"]}, {"_id": 0, "name": 1, "code": 1, "cash_balance": 1})
        if branch:
            item["branch_name"] = branch.get("name", "")
            item["branch_code"] = branch.get("code", "")
            item["cash_balance"] = branch.get("cash_balance", 0)
    
    # Calculate totals
    totals = {
        "total_sales": sum(r.get("total_sales", 0) for r in result),
        "total_profit": sum(r.get("total_profit", 0) for r in result),
        "total_transactions": sum(r.get("transactions", 0) for r in result)
    }
    
    return {
        "period": {"from": date_from, "to": date_to},
        "branches": result,
        "totals": totals
    }



# ==================== LAPORAN HUTANG (Payables) ====================

@router.get("/payables")
async def payables_report(
    supplier_id: str = "",
    status: str = "",  # unpaid, partial, paid
    user: dict = Depends(get_current_user)
):
    """Accounts Payable Report"""
    query = {}
    if supplier_id:
        query["supplier_id"] = supplier_id
    if status:
        if status == "unpaid":
            query["paid_amount"] = 0
        elif status == "partial":
            query["$expr"] = {"$and": [
                {"$gt": ["$paid_amount", 0]},
                {"$lt": ["$paid_amount", "$total"]}
            ]}
        elif status == "paid":
            query["$expr"] = {"$gte": ["$paid_amount", "$total"]}
    
    items = await purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    
    # Calculate totals
    total_payable = sum(max(0, item.get("total", 0) - item.get("paid_amount", 0)) for item in items)
    total_paid = sum(item.get("paid_amount", 0) for item in items)
    
    return {
        "items": items,
        "total_payable": total_payable,
        "total_paid": total_paid,
        "total_purchase": sum(item.get("total", 0) for item in items)
    }

# ==================== LAPORAN PIUTANG (Receivables) ====================

@router.get("/receivables")
async def receivables_report(
    customer_id: str = "",
    user: dict = Depends(get_current_user)
):
    """Accounts Receivable Report"""
    query = {"credit_balance": {"$gt": 0}}
    if customer_id:
        query["id"] = customer_id
    
    items = await customers.find(query, {"_id": 0}).sort("credit_balance", -1).to_list(1000)
    
    total_receivable = sum(item.get("credit_balance", 0) for item in items)
    
    return {
        "items": items,
        "total_receivable": total_receivable,
        "customer_count": len(items)
    }

# ==================== LAPORAN KAS (Cash Report) ====================

@router.get("/cash")
async def cash_report(
    date_from: str,
    date_to: str,
    branch_id: str = "",
    user: dict = Depends(get_current_user)
):
    """Cash flow report"""
    from database import db
    
    query = {
        "date": {"$gte": date_from, "$lte": date_to}
    }
    if branch_id:
        query["$or"] = [{"account_id": branch_id}, {"branch_id": branch_id}]
    
    cash_transactions_coll = db["cash_transactions"]
    items = await cash_transactions_coll.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    total_in = sum(t.get("amount", 0) for t in items if t.get("transaction_type") == "cash_in")
    total_out = sum(t.get("amount", 0) for t in items if t.get("transaction_type") == "cash_out")
    
    return {
        "period": {"from": date_from, "to": date_to},
        "items": items,
        "total_in": total_in,
        "total_out": total_out,
        "net_flow": total_in - total_out
    }

# ==================== LAPORAN SUPPLIER ====================

@router.get("/suppliers")
async def supplier_report(
    user: dict = Depends(get_current_user)
):
    """Supplier performance report"""
    # Get all suppliers with their purchase totals
    pipeline = [
        {"$match": {"status": {"$in": ["received", "partial"]}}},
        {
            "$group": {
                "_id": "$supplier_id",
                "supplier_name": {"$first": "$supplier_name"},
                "total_purchases": {"$sum": "$total"},
                "total_orders": {"$sum": 1},
                "total_paid": {"$sum": "$paid_amount"},
                "avg_order": {"$avg": "$total"}
            }
        },
        {"$sort": {"total_purchases": -1}}
    ]
    
    result = await purchase_orders.aggregate(pipeline).to_list(500)
    
    # Add supplier details
    for item in result:
        supplier = await suppliers.find_one({"id": item["_id"]}, {"_id": 0})
        if supplier:
            item["supplier_code"] = supplier.get("code", "")
            item["debt_balance"] = supplier.get("debt_balance", 0)
            item["phone"] = supplier.get("phone", "")
    
    return {
        "items": result,
        "total_suppliers": len(result),
        "total_purchases": sum(r.get("total_purchases", 0) for r in result)
    }

# ==================== LAPORAN BEST SELLERS ====================

@router.get("/best-sellers")
async def best_sellers_report(
    date_from: str,
    date_to: str,
    limit: int = 20,
    branch_id: str = "",
    user: dict = Depends(get_current_user)
):
    """Top selling products report"""
    match_query = {
        "status": "completed",
        "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
    }
    
    if branch_id:
        match_query["branch_id"] = branch_id
    
    pipeline = [
        {"$match": match_query},
        {"$unwind": "$items"},
        {
            "$group": {
                "_id": "$items.product_id",
                "product_name": {"$first": "$items.product_name"},
                "product_code": {"$first": "$items.product_code"},
                "quantity_sold": {"$sum": "$items.quantity"},
                "revenue": {"$sum": "$items.total"},
                "profit": {"$sum": {"$subtract": [
                    "$items.total",
                    {"$multiply": ["$items.quantity", "$items.cost_price"]}
                ]}}
            }
        },
        {"$sort": {"quantity_sold": -1}},
        {"$limit": limit}
    ]
    
    result = await transactions.aggregate(pipeline).to_list(limit)
    
    return {
        "period": {"from": date_from, "to": date_to},
        "items": result,
        "total_items": len(result)
    }

# ==================== LAPORAN KASIR ====================

@router.get("/cashiers")
async def cashier_report(
    date_from: str,
    date_to: str,
    branch_id: str = "",
    user: dict = Depends(get_current_user)
):
    """Cashier performance report"""
    from database import users
    
    match_query = {
        "status": "completed",
        "created_at": {"$gte": date_from, "$lte": date_to + "T23:59:59"}
    }
    
    if branch_id:
        match_query["branch_id"] = branch_id
    
    pipeline = [
        {"$match": match_query},
        {
            "$group": {
                "_id": "$cashier_id",
                "cashier_name": {"$first": "$cashier_name"},
                "total_sales": {"$sum": "$total"},
                "total_transactions": {"$sum": 1},
                "avg_transaction": {"$avg": "$total"},
                "total_profit": {"$sum": "$profit"}
            }
        },
        {"$sort": {"total_sales": -1}}
    ]
    
    result = await transactions.aggregate(pipeline).to_list(100)
    
    return {
        "period": {"from": date_from, "to": date_to},
        "items": result,
        "total_cashiers": len(result),
        "total_sales": sum(r.get("total_sales", 0) for r in result)
    }

# ==================== EXPORT EXCEL ====================

@router.get("/export/excel/{report_type}")
async def export_excel(
    report_type: str,
    date_from: str = "",
    date_to: str = "",
    branch_id: str = "",
    user: dict = Depends(get_current_user)
):
    """Export report to Excel"""
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io
    
    wb = Workbook()
    ws = wb.active
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    if report_type == "sales":
        ws.title = "Laporan Penjualan"
        # Get data
        data = await sales_report(date_from, date_to, branch_id, "day", user)
        
        headers = ["Tanggal", "Penjualan Bersih", "Laba", "Transaksi"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(data.get("data", []), 2):
            ws.cell(row=row_idx, column=1, value=item.get("_id")).border = border
            ws.cell(row=row_idx, column=2, value=item.get("net_sales", 0)).border = border
            ws.cell(row=row_idx, column=3, value=item.get("profit", 0)).border = border
            ws.cell(row=row_idx, column=4, value=item.get("transactions", 0)).border = border
    
    elif report_type == "inventory":
        ws.title = "Laporan Persediaan"
        data = await inventory_report(branch_id, "", False, user)
        
        headers = ["Kode", "Nama Produk", "Stok", "Nilai Modal", "Nilai Jual", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(data.get("data", []), 2):
            ws.cell(row=row_idx, column=1, value=item.get("product_code")).border = border
            ws.cell(row=row_idx, column=2, value=item.get("product_name")).border = border
            ws.cell(row=row_idx, column=3, value=item.get("quantity", 0)).border = border
            ws.cell(row=row_idx, column=4, value=item.get("stock_value", 0)).border = border
            ws.cell(row=row_idx, column=5, value=item.get("retail_value", 0)).border = border
            ws.cell(row=row_idx, column=6, value="Menipis" if item.get("is_low_stock") else "Aman").border = border
    
    elif report_type == "products":
        ws.title = "Laporan Produk"
        data = await product_performance_report(date_from, date_to, branch_id, 100, user)
        
        headers = ["Kode", "Nama Produk", "Qty Terjual", "Pendapatan", "Laba", "Margin %"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, item in enumerate(data.get("data", []), 2):
            ws.cell(row=row_idx, column=1, value=item.get("product_code")).border = border
            ws.cell(row=row_idx, column=2, value=item.get("product_name")).border = border
            ws.cell(row=row_idx, column=3, value=item.get("quantity_sold", 0)).border = border
            ws.cell(row=row_idx, column=4, value=item.get("net_revenue", 0)).border = border
            ws.cell(row=row_idx, column=5, value=item.get("profit", 0)).border = border
            ws.cell(row=row_idx, column=6, value=round(item.get("profit_margin", 0), 1)).border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"laporan_{report_type}_{date_from}_{date_to}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
